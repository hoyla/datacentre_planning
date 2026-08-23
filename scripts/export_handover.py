"""Generate the data-team handover workbook.

Six sheets, in the order a new reader meets them:

- **Read me** — one row per column: what it contains, how it was derived,
  what its caveats mean. At this width the dictionary is load-bearing,
  not a nicety.
- **Sites** — one row per site, wide deliberately: the sheet exists so a
  reporter can sort by one dimension and read the others on the same row.
  Grouped left-to-right: identity & links, status & recency, power,
  environment, parties, findings coverage, nearest energy project,
  coverage & gaps, Barbour.
- **Applications** — one row per member application, linked by site_key.
- **Energy projects** — the NSIP energy layer: nationally significant
  generation/transmission projects, ranked by distance to the nearest
  data-centre site. A different unit of analysis, kept in the same
  workbook because the site rows point into it.
- **External aggregates** — regulators' and network operators' published
  aggregate figures on data centre power demand, presented beside — and
  deliberately never joined to — the planning-derived rows. They measure
  different quantities from planning documents, which is itself the
  point: the sheet shows what each side can and cannot see.
- **Capacity claims** — site-level figures from named external sources
  (currently NESO's Existing Agreements Register), one row per claim as
  the source states it. Where a claim is matched to a site the match is
  a hand-adjudicated inference and its confidence, method and written
  evidence are columns. The Sites sheet's power columns remain
  planning-derived only; a register figure never becomes a site's
  number.
- **Provenance** — run metadata plus the corpus-level statement of known
  retrieval gaps, so the coverage caveats live in the deliverable rather
  than in a covering email.

Sites without planning applications are still rows: 55 Barbour-recorded
projects at pre-planning stage have no public material anywhere, and a
workbook that omitted them would silently equate "no application yet"
with "no site". Their coverage columns say exactly what they are.

Every heuristic column carries its source; nothing here is
hand-maintained — regenerate after any pipeline change and the workbook
reflects the database.

The workbook is an *interface*, not a store: annotations belong in the
designated annotation tab of the shared copy (or a separate sheet),
never in generated columns.

Usage:
    .venv/bin/python scripts/export_handover.py
        [--out data/exports/dc_build_handover_<date>.xlsx]
"""

from __future__ import annotations

import argparse
import datetime as dt
import json
import re
import subprocess
import sys
from math import asin, cos, radians, sin, sqrt
from pathlib import Path, PurePosixPath

sys.path.insert(0, str(Path(__file__).parent.parent))

from dotenv import load_dotenv

load_dotenv(Path(__file__).parent.parent / ".env")

from dcp import adjudication_gate  # noqa: E402
from dcp import db  # noqa: E402


SITE_SQL = """
WITH latest AS (
  SELECT DISTINCT ON (application_id) application_id, verdict, confidence, model
  FROM triage ORDER BY application_id, inserted_at DESC, id DESC),
app_docs AS (
  SELECT application_id, count(*) AS n FROM documents GROUP BY application_id),
app_findings AS (
  -- signal_family, not signal_type: the extraction prompt asks the model
  -- to name what it found in its own words, which produced 54,044
  -- distinct labels. Aggregating those per application yields a
  -- spreadsheet cell nobody can read or filter. The family is the
  -- 25-value canonical index over them (dcp/signal_families.py); the
  -- original label is still on every findings row for anyone drilling in.
  --
  -- count(DISTINCT passage), not count(*). Several models read the same
  -- documents, so one sentence found by three readers was three rows and
  -- the column read as three findings. Measured 2026-08-10: 1,019,106
  -- rows over 878,651 distinct passages -- 14% inflation, smaller than
  -- feared, but a count should mean one thing and "three readers agreed"
  -- is corroboration rather than volume.
  SELECT application_id,
         count(DISTINCT (document_id, md5(evidence_text), evidence_page))
                                                               AS n,
         array_agg(DISTINCT signal_family) FILTER (
             WHERE signal_family IS NOT NULL
               AND signal_family <> 'unclassified')            AS signal_families
  FROM findings GROUP BY application_id),
app_power AS (
  -- Adjudicated capacity ONLY. The previous version of this query took
  -- max(value_number) over every finding carrying a 'MW' unit, which is
  -- wrong in a way that reaches the reader: planning statements argue for
  -- approval by quoting market forecasts, policy targets and other
  -- schemes, so the largest MW figure in a site's documents is usually
  -- not about that site. Under that rule a Slough application reported
  -- 30GW (an NGESO storage target) and a Chiltern one 22,700MW (a Savills
  -- market forecast).
  --
  -- power_adjudication resolves whose figure each one is; only
  -- verdict='site_capacity' is admitted here. The quantities stay in
  -- separate columns because IT load, grid connection and standby
  -- generation are different numbers for the same site — medians across
  -- the corpus are 44 MW, 99 MW and 3.3 MW — and a single "site MW"
  -- column would silently mix them.
  SELECT application_id,
         max(value_mw) FILTER (WHERE quantity_type = 'it_load')     AS it_load_mw,
         max(value_mw) FILTER (WHERE quantity_type = 'total_site')  AS total_site_mw,
         max(value_mw) FILTER (WHERE quantity_type = 'grid_connection')
                                                                    AS grid_mw,
         max(value_mw) FILTER (WHERE quantity_type = 'onsite_generation')
                                                                    AS gen_mw,
         count(*)                                                   AS n_capacity,
         count(*) FILTER (WHERE is_maximum)                         AS n_ultimate
  FROM power_adjudication
  WHERE verdict = 'site_capacity' AND value_mw IS NOT NULL
  GROUP BY application_id),
app_power_excluded AS (
  -- Recorded so a reader can see that figures were considered and set
  -- aside, rather than wondering why a number in the documents is absent
  -- from the workbook.
  SELECT application_id, count(*) AS n_excluded
  FROM power_adjudication WHERE verdict <> 'site_capacity'
  GROUP BY application_id),
app_provenance AS (
  -- How this application's documents were obtained. Hand-ingested
  -- documents carry file:// URIs (no portal link to offer a reader);
  -- exports must label them rather than leave an apparent gap.
  SELECT application_id,
         count(*) FILTER (WHERE url LIKE 'file://%%') AS manual_docs,
         count(*) FILTER (WHERE url NOT LIKE 'file://%%') AS portal_docs
  FROM documents GROUP BY application_id),
app_eia AS (
  -- Heuristic EIA indicators, two independent signals: the application
  -- reference carries an EIA-shaped suffix (FULEA/OUTES/EIASR/SCR/SCO/
  -- SCREEN), or the held documents include Environmental Statement
  -- material. Authoritative status (screening outcomes) comes from
  -- deep-read; the column is deliberately labelled heuristic.
  SELECT a.id AS application_id,
         (a.application_ref ~* '(EA|ES)$|EIASR|/SCR|/SCO|SCREEN') AS ref_hit,
         EXISTS (SELECT 1 FROM documents d WHERE d.application_id = a.id
                 AND (d.url ~* 'ENVIRONMENTAL_STATEMENT|ES_APPENDIX|ENV_STATEMENT'
                      OR d.kind ~* 'environmental statement')) AS doc_hit
  FROM applications a)
SELECT s.site_key, s.classification, s.display_name,
       s.latitude, s.longitude, s.coord_source,
       array_agg(DISTINCT split_part(a.application_ref,'/',1))
           FILTER (WHERE a.id IS NOT NULL)                    AS councils,
       count(DISTINCT a.id)                                   AS n_apps,
       array_agg(DISTINCT a.application_ref)
           FILTER (WHERE a.id IS NOT NULL)                    AS app_refs,
       array_agg(DISTINCT coalesce(l.verdict,'?'))
           FILTER (WHERE a.id IS NOT NULL)                    AS verdicts,
       coalesce(sum(ad.n), 0)                                 AS docs_held,
       coalesce(sum(af.n), 0)                                 AS findings_n,
       max(pw.it_load_mw)                                     AS it_load_mw,
       max(pw.total_site_mw)                                  AS total_site_mw,
       max(pw.grid_mw)                                        AS grid_mw,
       max(pw.gen_mw)                                         AS gen_mw,
       coalesce(sum(pw.n_capacity), 0)                        AS n_capacity,
       coalesce(sum(px.n_excluded), 0)                        AS n_excluded,
       -- Ranked by evidence volume, not listed alphabetically. A flat
       -- presence list marks nearly every family present on any
       -- document-heavy site, so it ends up reporting "this site has a
       -- lot of documents" rather than what the site is about. Counts
       -- ordered by size discriminate: a site whose largest families are
       -- power_generation and power_grid reads differently from one
       -- dominated by ecology_biodiversity and designated_sites.
       --
       -- Correlated rather than joined because array_agg cannot flatten
       -- the per-application arrays across a site group; runs once per
       -- site against an indexed column.
       (SELECT array_agg(fam || ' (' || n || ')' ORDER BY n DESC, fam)
          FROM (SELECT f2.signal_family AS fam, count(*) AS n
                  FROM findings f2
                  JOIN site_members m3
                       ON m3.application_id = f2.application_id
                       AND m3.retired_at IS NULL
                 WHERE m3.site_id = s.id
                   AND f2.signal_family IS NOT NULL
                   AND f2.signal_family <> 'unclassified'
                 GROUP BY f2.signal_family) ranked)         AS signal_families,
       bool_or(ae.ref_hit)                                    AS eia_ref_hit,
       bool_or(ae.doc_hit)                                    AS eia_doc_hit,
       coalesce(sum(ap.manual_docs), 0)                       AS manual_docs,
       (SELECT p2.external_ref FROM site_members m2
          JOIN projects p2 ON p2.id = m2.project_id
        WHERE m2.site_id = s.id AND m2.retired_at IS NULL
        ORDER BY p2.external_ref LIMIT 1)                     AS ptno,
       max(p.title)                                           AS barbour_title,
       max(p.stage_summary)                                   AS barbour_stage,
       max(p.value_gbp)                                       AS barbour_value_gbp,
       max(p.floor_area)                                      AS barbour_floor_area,
       max(p.site_area)                                       AS barbour_site_area,
       max(p.plan_date)                                       AS barbour_plan_date,
       max(p.decision_date)                                   AS barbour_decision_date,
       -- For the DESNZ consumption context: a Barbour-anchored site with
       -- no applications has no council prefixes, and a site spanning
       -- councils needs Barbour's authority to say which one it is in.
       max(p.authority_name)                                  AS barbour_authority
FROM sites s
LEFT JOIN site_members m ON m.site_id = s.id AND m.retired_at IS NULL
LEFT JOIN applications a ON a.id = m.application_id
LEFT JOIN latest l ON l.application_id = a.id
LEFT JOIN app_docs ad ON ad.application_id = a.id
LEFT JOIN app_findings af ON af.application_id = a.id
LEFT JOIN app_power pw ON pw.application_id = a.id
LEFT JOIN app_power_excluded px ON px.application_id = a.id
LEFT JOIN app_eia ae ON ae.application_id = a.id
LEFT JOIN app_provenance ap ON ap.application_id = a.id
LEFT JOIN projects p ON p.id = m.project_id
WHERE s.retired_at IS NULL
GROUP BY s.id
ORDER BY s.site_key
"""

APP_SQL = """
WITH latest AS (
  SELECT DISTINCT ON (application_id) application_id, verdict, confidence,
         model, why, signals
  FROM triage ORDER BY application_id, inserted_at DESC, id DESC),
app_docs AS (
  SELECT application_id, count(*) AS n FROM documents GROUP BY application_id),
app_findings AS (
  -- Distinct passages, matching the Sites query above: one sentence found
  -- by three readers is one finding corroborated three times, not three.
  SELECT application_id,
         count(DISTINCT (document_id, md5(evidence_text), evidence_page)) AS n
  FROM findings GROUP BY application_id)
SELECT s.site_key, a.application_ref, m.joined_via,
       split_part(a.application_ref,'/',1) AS council,
       a.status, a.date_received, a.date_decided,
       l.verdict, l.confidence, l.model, l.why,
       array_to_string(l.signals, ', ') AS signals,
       a.url, coalesce(ad.n,0) AS docs_held, coalesce(af.n,0) AS findings_n,
       a.address, a.description
FROM sites s
JOIN site_members m ON m.site_id = s.id AND m.retired_at IS NULL
JOIN applications a ON a.id = m.application_id
LEFT JOIN latest l ON l.application_id = a.id
LEFT JOIN app_docs ad ON ad.application_id = a.id
LEFT JOIN app_findings af ON af.application_id = a.id
WHERE s.retired_at IS NULL
ORDER BY s.site_key, a.application_ref
"""

# Barbour-recorded projects with no linked planning application: the
# pre-planning pipeline. Site materialisation already creates rows for
# most of these (29 of 55 at time of writing, membership via project_id),
# so this query is a safety net, not the primary path: anything the
# sites table already carries is skipped at append time, and only the
# genuinely absent are added. Everything unknowable about them is
# honestly blank, and the coverage columns say why.
BARBOUR_ONLY_SQL = """
SELECT p.external_ref, p.title, p.stage_summary, p.dev_type,
       p.authority_name, p.address, p.description,
       p.latitude, p.longitude,
       p.value_gbp, p.floor_area, p.site_area,
       p.plan_date, p.decision_date,
       -- The role blocks. For a pre-planning row Barbour is the only
       -- source there is, so leaving its parties blank would empty the
       -- column exactly where it is the whole of what is known.
       p.raw_metadata
FROM projects p
WHERE NOT EXISTS (SELECT 1 FROM project_applications pa
                  WHERE pa.project_id = p.id)
ORDER BY p.external_ref
"""

# The NSIP energy layer: metadata only, by design — a single DCO document
# set runs to thousands of files, and the project page already states
# what the layer needs (name, description, often capacity).
NSIP_SQL = """
SELECT a.application_ref, a.status,
       (a.raw_metadata->>'location_y')::float  AS lat,
       (a.raw_metadata->>'location_x')::float  AS lon,
       a.raw_metadata->'pins_page'->>'name'    AS name,
       coalesce(a.raw_metadata->'other_fields'->>'applicant_name',
                a.raw_metadata->'pins_page'->>'applicant') AS applicant,
       a.raw_metadata->>'app_type'             AS app_type,
       a.raw_metadata->>'area_name'            AS region,
       a.raw_metadata->'pins_page'->>'stage'   AS stage,
       a.raw_metadata->'pins_page'->'capacity_mentions' AS capacity,
       a.raw_metadata->'pins_page'->>'description'      AS description,
       a.raw_metadata->'pins_page'->>'developer_site'   AS developer_site,
       a.raw_metadata->'other_fields'->>'date_application_accepted' AS accepted,
       a.url
FROM applications a
WHERE a.discovered_via @> ARRAY['nsip_energy']
  AND (a.raw_metadata->>'location_x') ~ '^-?[0-9.]+$'
ORDER BY a.application_ref
"""

# How many finding families to name per site before summarising the rest.
# Six covers what characterises a site without turning the cell into a
# list of everything the corpus can detect.
TOP_FAMILIES = 6

# Statuses that mean a decision has effectively been reached even where
# no decision date was recorded. Used only for the liveness rollup, whose
# operational definition the dictionary states verbatim.
_DECIDED_STATUS_RE = re.compile(
    r"withdraw|decided|granted|refus|approv|permit|reject|dismiss|"
    r"determin|disposed", re.I)

DRIVE_LEDGER = Path("data/exports/.drive_sync_state.json")

SITE_HEADERS = [
    # --- identity & links -------------------------------------------------
    "Site key", "Classification", "Site name", "Proposal",
    "Proposal describes a development?",
    "Latitude", "Longitude", "Coordinate source", "Councils",
    # The workbook is the interface to the Drive-by-site archive and to
    # the public record; a row that cannot reach its evidence is a claim,
    # not an index card.
    "Drive folder", "Source portal",
    # --- status & recency -------------------------------------------------
    "Applications", "Application refs",
    "Applications without recorded decision",
    "Latest application date", "Latest decision date",
    "Verdict mix (v1 triage)",
    # --- power --------------------------------------------------------
    # The ranking column comes first and is always a number where any
    # basis exists, because capacity is how these sites get compared. Its
    # qualifications sit immediately to its right rather than in a
    # methodology note: whoever sorts by MW sees, in the adjacent cell,
    # whether they are sorting a disclosed figure or an inference.
    "Power MW (best available)", "Power basis", "Power confidence",
    "Power caveat", "Figures provisional?",
    # The components stay, because they are different quantities rather
    # than competing estimates and some questions need them separately.
    "IT load MW (adjudicated)", "Total site MW (adjudicated)",
    "Grid connection MW (adjudicated)", "On-site generation MW (adjudicated)",
    # What the generation figure is, by the passage that states it: one
    # machine's rating or the figure as stated — and the largest fleet
    # the documents disclose by count and rating, never multiplied. Three
    # columns because they are three different things (dcp/site_profile).
    "On-site generation figure basis", "Generator units disclosed (count)",
    "Generator unit rating MW (disclosed)",
    "Capacity figures attributed to site", "Power figures excluded (context)",
    # --- environment ----------------------------------------------------
    "Facility character", "Scale band", "Scale basis",
    # From dcp/site_profile, shared with the web view so both present the
    # same signal for the same reason.
    "Standby generators (count)", "Generation type", "Generator caveat",
    # Cooling method rather than a water-consumption figure: the corpus
    # supports the first and not the second, and cooling method is what
    # actually determines water demand.
    "Cooling method", "Water evidence", "Cooling caveat",
    "EIA status (from documents)", "EIA indicators (heuristic)",
    "Environmental subjects (description keywords)",
    # --- parties ----------------------------------------------------------
    # Who is behind the scheme, from the source that states it rather
    # than from the one that can only count it (dcp/site_profile). One
    # column per kind of claim: what Barbour records as the end user and
    # the client; the group a person has confirmed that name belongs to,
    # beside it and never in place of it; the advisers Barbour names;
    # and, kept separate, the organisations the documents name with the
    # count that is all that claim rests on. Every party of every role,
    # including the roles too specific for a site row, is a row of the
    # Parties sheet.
    "End user (Barbour)", "Applicant of record (Barbour)",
    "Operator group (confirmed alias)", "Advisers (Barbour)",
    "Also named in documents (mention counts)",
    "Planning authority (register)", "Parties source",
    # --- findings coverage -------------------------------------------------
    "Finding subjects (top families by volume)",
    "Documents held", "Documents analysed", "Verified findings",
    "Documents obtained by hand",
    # --- nearest energy project ---------------------------------------------
    "Nearest energy project (NSIP)", "Distance to energy project (km)",
    "Energy project stated capacity",
    # --- consumption context (DESNZ) ----------------------------------------
    # Context, never attribution: the authority's large-user consumption
    # change beside the national one, with the inferred authority named
    # so the mapping is visible beside its product.
    "Local authority (DESNZ series, inferred)",
    "Local authority large-user electricity 2019→2024 (% change)",
    "National large-user electricity 2019→2024 (% change)",
    # --- coverage & gaps ---------------------------------------------------
    "Capacity status", "Acquisition status",
    # --- Barbour -----------------------------------------------------------
    "Barbour Ptno", "Barbour title",
    "Barbour stage", "Barbour value £", "Barbour floor area sqm",
    "Barbour site area", "Barbour plan date", "Barbour decision date",
]

APP_HEADERS = [
    "Site key", "Application ref", "Joined site via", "Council", "Status",
    "Date received", "Date decided", "Verdict (latest)", "Verdict confidence",
    "Verdict model", "Verdict reasoning", "Signals", "Portal URL",
    "Drive folder", "Documents held", "Verified findings",
    "Environmental signals (description keywords)", "Address", "Description",
]

ENERGY_HEADERS = [
    "Project reference", "Project name", "Stated capacity",
    "Type of application", "Stage", "Status", "Applicant", "Region",
    "Date accepted", "Nearest data-centre site", "Distance (km)",
    "PINS project page", "Developer website", "Description",
]

# One row per column. This is journalist-facing text: it states what the
# data contains and how it was derived, never how to use it.
DICTIONARY: list[tuple[str, str, str]] = [
    ("Sites", "Site key",
     "Stable identifier for the site; also the prefix of its Drive folder "
     "name. Keys beginning 'ptno-' are anchored on a Barbour project."),
    ("Sites", "Classification",
     "How the site entered the universe, and therefore what kind of "
     "evidence stands behind it. Five values, from dcp/sites.py: "
     "**both** — the cluster holds at least one planning application in "
     "the data-centre universe AND a Barbour ABI project record, so two "
     "independent sources agree the site exists. "
     "**ours_only** — planning applications we found and classified, with "
     "no matching Barbour project. Those applications arrived by several "
     "routes, not one: the national keyword sweep, spatial searches "
     "around known sites, operator and agent name searches, the energy "
     "adjacency sweep, the Foxglove list and parent-application backfill. "
     "The 'How we found it' column names the routes for each site. "
     "**unlocatable** — the same as ours_only except that not one of its "
     "applications carries a coordinate, so the site is real but cannot "
     "be placed on a map. These are the sites absent from the map view. "
     "**barbour_covered** — anchored on a Barbour project, with planning "
     "applications in the cluster that our own classification did not put "
     "in the universe. "
     "**barbour_only** — a Barbour project with no planning application "
     "at all: the scheme is known at pre-planning stage and has not yet "
     "reached a public register."),
    ("Sites", "Site name",
     "Display name assembled at site materialisation; for pre-planning "
     "rows, the Barbour project title."),
    ("Sites", "Proposal",
     "A one-line description of what is proposed, taken word-for-word from "
     "the council's own application description. Planning descriptions "
     "usually open with procedure — which condition is being discharged, "
     "which permission varied — and state the development somewhere in the "
     "middle; this is the clause that describes the development, selected "
     "across all of the site's applications. It is an extract, never a "
     "paraphrase, so it can be quoted; the untouched description is on "
     "every Applications row."),
    ("Sites", "Proposal describes a development?",
     "'No' where nothing on the public record describes what is being "
     "built — the site is known only through condition discharges, "
     "consultations or screening requests — and the Proposal cell is "
     "therefore procedural text rather than a summary."),
    ("Sites", "Latitude / Longitude / Coordinate source",
     "Best available coordinates and where they came from. Rows without "
     "coordinates cannot be matched in the distance columns."),
    ("Sites", "Councils",
     "Council prefixes of the member applications; for pre-planning rows, "
     "the authority Barbour records."),
    ("Sites", "Drive folder",
     "Link to this site's folder of source documents on the shared Drive. "
     "Blank means the site has not been synced yet (see Acquisition "
     "status), not that documents are missing."),
    ("Sites", "Source portal",
     "Link to one of the site's applications on the council's public "
     "register — the most recently received one holding a link. Per-"
     "application links are on the Applications sheet. Where the label "
     "reads '1 of N registers', the site spans more than one planning "
     "authority and each keeps its own register: the link reaches one of "
     "them, and the Applications sheet has the rest."),
    ("Sites", "Applications / Application refs",
     "Member planning applications of this site (grouping decided at "
     "materialisation; 'Joined site via' on the Applications sheet says "
     "how each joined)."),
    ("Sites", "Applications without recorded decision",
     "Count of member applications with no decision date recorded and no "
     "decided-type status text at last fetch. An approximation of 'live': "
     "councils record decisions unevenly, so treat as an indicator, not "
     "a legal status."),
    ("Sites", "Latest application date / Latest decision date",
     "Most recent received date and most recent decision date across the "
     "member applications, as recorded by the council."),
    ("Sites", "Verdict mix (v1 triage)",
     "Distinct triage verdicts across member applications (v1 rubric)."),
    ("Sites", "Power MW (best available)",
     "One rankable capacity figure per site. Falls back through disclosed "
     "IT load, total site power, grid connection, on-site generation, "
     "then a floor-area inference — losing authority at each step. The "
     "three columns to its right say which step and how much to trust it."),
    ("Sites", "External power indicators",
     "Whether this site has a live match in the Capacity claims sheet, "
     "and the strongest confidence tier among its matches: 'strong', "
     "'probable' or 'tentative', with a count where there is more than "
     "one. Deliberately not a megawatt figure — a claim can be a "
     "different quantity from this row's own Power MW (a contracted "
     "grid ceiling is not IT load), so a number here would read as "
     "directly comparable to Power MW when it is not. Hover the tag, or "
     "open the site panel, for the claim names and their own figures."),
    ("Sites", "Where the power columns come from, and what not to compare",
     "Every power column on this sheet is read from the site's own "
     "planning documents and adjudicated as describing this site — the "
     "applicant's stated figures, not anyone else's estimate of them. "
     "The Capacity claims sheet holds a different class of number "
     "entirely: grid registers, accounts filed at Companies House, and "
     "operators' own websites, each measuring a different quantity with "
     "different authority behind it. The two are deliberately kept in "
     "separate sheets and must not be compared cell against cell. Where "
     "they diverge — an operator's website claiming more than its own "
     "audited accounts, a contracted grid ceiling far above a declared "
     "IT load — the divergence is a finding to report, not an error to "
     "reconcile."),
    ("Sites", "Figures provisional?",
     "Whether this row's findings-derived values come from a complete "
     "reading. Where they do not, every such value is a floor rather than a "
     "measurement: the largest capacity in the documents read so far, the "
     "generators counted so far, the applicants named so far. Further "
     "reading can raise these figures and cannot lower them. A campus "
     "promoted as 1GW may show 500MW here simply because the document "
     "stating the larger figure has not been analysed yet."),
    ("Sites", "Power basis / Power confidence / Power caveat",
     "What the figure is based on, how strong that basis is, and any "
     "qualification. These qualify the Power MW column and travel with "
     "it: a sort by MW should be read with all four columns."),
    ("Sites", "IT load / Total site / Grid connection / On-site generation MW",
     "The adjudicated components, kept separate because they are "
     "different quantities for the same site, not competing estimates. "
     "IT load is what the racks draw and excludes cooling overhead; total "
     "site includes it, so the two are not comparable and a site quoting "
     "only one is not smaller than a site quoting the other. Grid "
     "connection is capacity sought, reserved or contracted — headroom, "
     "which operators commonly secure more of than they draw, and which "
     "phased schemes take up over years. On-site generation is standby "
     "and CHP plant; read it beside IT load rather than alone, because "
     "generation far below load usually means life-safety backup only and "
     "a wholly grid-dependent site, which is itself worth reporting."),
    ("Sites", "What is NOT in the generation and capacity columns",
     "Three quantities that look like power and are deliberately kept "
     "out of these columns, each recorded under its own type so it stays "
     "findable: battery and UPS ratings (energy_storage) state discharge "
     "speed, not generation or demand; thermal input (thermal_input) is "
     "fuel entering a plant, typically two to three times the electricity "
     "leaving it; and annual energy consumption, which is not a capacity "
     "at all — one application states a load in kW that is really a "
     "year's kWh, and taken literally implies a site four times the "
     "national grid. Figures above 3 GW are rejected on that basis alone."),
    ("Sites", "How much to trust a single capacity figure",
     "Some sites state a figure once and nothing corroborates it; others "
     "state it and their grid connection or standby plant independently "
     "agrees. Both appear in this column identically, so where it "
     "matters, check the site panel: it shows the components beside each "
     "other. Two known limits travel with these figures. A figure the "
     "documents describe as per-building is not the site total — one "
     "scheme states 75 MW per building and, elsewhere, three buildings — "
     "and a generation figure taken from one machine's specification is "
     "not the fleet, where the documents describe dozens of units."),
    ("Sites", "On-site generation figure basis",
     "What the 'On-site generation MW' figure is, read from the passages "
     "that state it: **per unit** — the documents present it as one "
     "machine's rating (\"112 No. standby generators (likely to be "
     "3.2MWe\"), so it is not the site's generation; **as stated** — the "
     "figure stands as the documents give it. Deterministic (dcp/"
     "site_profile.generation_figure): it fires only where a passage "
     "states a count or 'each' and the rating it gives matches the "
     "stored figure, and where a stated count times a stated rating "
     "equals the figure, the figure is that total whatever another "
     "sentence calls it. Blank where the site has no generation figure. "
     "A per-row adjudication of every generation figure is planned and "
     "will supersede this label."),
    ("Sites", "Generator units disclosed (count) / Generator unit rating MW (disclosed)",
     "The largest fleet of generators on this site that the documents "
     "disclose as a count and a rating in one passage (\"up to 650 no. "
     "2,480 kW back-up diesel generators\"), "
     "reported beside the generation figure and never multiplied into it. "
     "Where the generation figure is itself a per-unit rating, these are "
     "that rating and its count. A count here is what one passage says; "
     "the 'Standby generators (count)' column is the highest count in any "
     "document and may differ."),
    ("Sites", "Capacity figures attributed to site",
     "How many megawatt figures the adjudication attributed to this site "
     "itself."),
    ("Sites", "Power figures excluded (context)",
     "Megawatt figures found in this site's documents but judged to be "
     "market context (forecasts, policy targets, other schemes) — "
     "considered and set aside, not missed."),
    ("Sites", "Facility character",
     "What kind of facility the descriptions indicate (rule-based over "
     "application descriptions)."),
    ("Sites", "Scale band / Scale basis",
     "Size band implied by the power figure, and whether that figure was "
     "stated in documents or inferred from floor area at 1.71 kW per sqm "
     "(measured across the 53 sites disclosing both)."),
    ("Sites", "Standby generators (count)",
     "Plant: the highest generator count disclosed in any one of this "
     "site's documents, so separately-described phases are not summed. "
     "The only count on these columns that is a quantity of equipment — "
     "the bracketed numbers in 'Generation type' and 'Cooling method' are "
     "mentions, and the two do not reconcile. Not adjudicated (unlike "
     "capacity): generator counts are rarely quoted as market context, "
     "which capacity routinely is."),
    ("Sites", "Generation type",
     "Fuels named in generation-related findings, counted by how many "
     "passages name each — not by how many generators run on each, which "
     "the documents do not break down. Fuels well below the leader are "
     "listed as 'also referenced' — usually options weighed rather than "
     "plant installed. 'CHP' flags combined heat and power language."),
    ("Sites", "Generator caveat",
     "Standing qualification on the generator columns."),
    ("Sites", "Cooling method",
     "Cooling technologies named in the documents, counted by how many "
     "passages name each rather than by how much plant is installed. "
     "Applications routinely compare options before choosing, so more than "
     "one may appear; the count separates the method used from the methods "
     "considered. Cooling method is reported instead of a water-consumption "
     "figure because it is what determines water demand and because the "
     "documents support it: an air-cooled hall and an evaporative one differ "
     "by orders of magnitude."),
    ("Sites", "Water evidence",
     "How much the documents say about water consumption or abstraction, as "
     "a count of findings — not a volume. The water and cooling finding "
     "families are large but dominated by flood and drainage engineering "
     "that every development produces (rainfall depths, pipe runs, design "
     "discharge rates). Filtered to consumption and abstraction, only 119 of "
     "429 sites disclose anything at all. That silence is itself a finding: "
     "no volume is published here because the applications do not contain "
     "one."),
    ("Sites", "EIA status (from documents)",
     "Environmental Impact Assessment status stated in the documents "
     "themselves (screening/scoping/ES submitted), precedence-ordered so "
     "an outcome beats a process step."),
    ("Sites", "EIA indicators (heuristic)",
     "Weaker, independent indicators: an EIA-shaped application reference "
     "or Environmental Statement material among held documents. Kept "
     "separate from the document-derived status deliberately."),
    ("Sites", "Environmental subjects (description keywords)",
     "Deterministic keyword extraction from application descriptions — a "
     "floor, not a census; the substantive content lives in documents."),
    ("Sites", "End user (Barbour); Applicant of record (Barbour); "
              "Advisers (Barbour)",
     "Organisations as Barbour ABI's project record states them, names "
     "only — the same records carry named individuals and their contact "
     "details, which stay in the source data and are not exported. End "
     "user is the party the scheme is for; applicant of record is "
     "Barbour's client; advisers are its planner, agent, architect and "
     "M&E engineer. Every other role Barbour records is a row on the "
     "Parties sheet."),
    ("Sites", "Operator group (confirmed alias)",
     "The corporate group a name belongs to, where a person has "
     "confirmed the link with evidence in "
     "data/priors/organisation_aliases.yaml — 'Ark Estates 5 Ltd' is "
     "Ark Data Centres. Sits beside the raw name and never replaces it; "
     "empty where no confirmed group covers the name. Names are grouped "
     "by evidence, never by resemblance: near-identical company names in "
     "this sector are routinely different companies."),
    ("Sites", "Also named in documents (mention counts)",
     "Organisations the site's documents name, and how many times. A "
     "count of mentions is not a role: the firm that wrote the planning "
     "statement is named more often than the developer, and a utilities "
     "section names whoever has ducts in the road. Organisations named "
     "once are not listed, and the Parties sheet records how many were "
     "dropped per site."),
    ("Sites", "Planning authority (register)",
     "The council whose register the site's applications sit in, from "
     "the application record — not from an organisation named in the "
     "documents. For a pre-planning project with no application, "
     "Barbour's authority."),
    ("Sites", "Parties source",
     "Which of the two sources this row's parties came from: the Barbour "
     "project record, the documents, both, or neither."),
    ("Parties", "site_key; role; organisation; source; source ref",
     "One row per organisation per role per site — the long form the "
     "Sites columns summarise. 'role' is end user, applicant, operator, "
     "adviser, other (any further Barbour role) or named in documents. "
     "'source' is barbour or documents; 'source ref' is the Barbour "
     "project number or the mention count. 'organisation' is the raw "
     "name as its source writes it and is never rewritten."),
    ("Parties", "group; Barbour role",
     "The confirmed alias group beside the raw name, empty where there "
     "is none; and, for a Barbour row, the role exactly as Barbour "
     "writes it, which is more specific than the role column."),
    ("Sites", "Finding subjects (top families by volume)",
     "The site's largest finding families by count — what its documents "
     "are substantively about. Counts are evidence volume, not "
     "conclusions."),
    ("Sites", "Documents held / Documents analysed",
     "Documents fetched for this site, and how many of them the deep-read "
     "has successfully analysed. Unanalysed documents can still contain "
     "anything; see Capacity status."),
    ("Sites", "Verified findings",
     "Deep-read findings whose evidence quotes passed verbatim "
     "verification against the source text before storage. Counted as "
     "distinct passages, not rows: several models read these documents, "
     "and one sentence found independently by three readers is one "
     "finding corroborated three times rather than three findings. "
     "Across the corpus that distinction removes about 16% of the raw "
     "row count. A high number here means a document-rich site, not "
     "necessarily an information-rich one — a long environmental "
     "statement yields hundreds of findings about drainage."),
    ("Sites", "Documents obtained by hand",
     "Documents ingested manually rather than fetched from a portal; they "
     "carry no public link."),
    ("Sites", "Nearest energy project (NSIP) / Distance (km) / stated capacity",
     "The nearest nationally significant energy project (see Energy "
     "projects sheet), straight-line distance, and any capacity its PINS "
     "page states. Shown for every site with coordinates regardless of "
     "distance — a large distance is itself information. Distance is to "
     "the nearest *located* site; sites without coordinates cannot match."),
    ("Sites", "Local authority (DESNZ series, inferred)",
     "The current local authority this site's council references — or, "
     "for pre-planning rows, its Barbour-recorded authority — resolve "
     "to, named so the two consumption columns beside it say whose "
     "figures they are. The match is an inference; the Councils column "
     "keeps the original values. Blank where the site's references span "
     "more than one authority and nothing selects among them, where the "
     "authority is outside Great Britain (the DESNZ series does not "
     "cover Northern Ireland), or where the planning authority is a "
     "development corporation rather than a local authority."),
    ("Sites", "Local authority / National large-user electricity "
     "2019→2024 (% change)",
     "Change in Half-Hourly-metered non-domestic electricity consumption "
     "between 2019 and 2024, for the site's local authority and for "
     "Great Britain, from DESNZ sub-national electricity statistics "
     "(Open Government Licence v3). Large users — half-hourly-metered "
     "non-domestic consumers — are the class data centres belong to, and "
     "DESNZ publishes it at local-authority level only: every per-MSOA "
     "row in the source carries zero half-hourly meters, so nothing "
     "finer exists. The figure describes the authority, not the site: an "
     "authority's total covers all its large users. The series ends "
     "2024, so later energisations are not in it, and authority figures "
     "are floors — DESNZ could not allocate a national remainder "
     "(~2.9 TWh in 2024) to any authority."),
    ("Sites", "Capacity status",
     "What the power columns' content (or emptiness) means: disclosed / "
     "inferred from floor area / analysed in full with nothing disclosed "
     "/ documents not yet analysed / no documents held / pre-application. "
     "These are different facts and the distinction is the point."),
    ("Sites", "Acquisition status",
     "Whether source material could be, and was, retrieved. Records the "
     "two known blocks explicitly: Coventry's register rejects "
     "non-browser clients (bot protection, not worked around), and one "
     "Wiltshire entry requires a consultee login."),
    ("Sites", "Barbour columns",
     "Barbour ABI project intelligence (licensed; credit required in "
     "published output). The title often carries a promoter's capacity "
     "claim — deliberately never copied into the Power MW column: "
     "promoter names routinely overstate what documents later disclose."),
    ("Applications", "Verdict (latest) / confidence / model / reasoning",
     "Latest triage verdict for the application, with the model's stated "
     "reasoning. Earlier verdicts are retained in the database."),
    ("Applications", "Portal URL",
     "The application's page on the council's public register."),
    ("Applications", "Drive folder",
     "Link to this application's own folder in the Drive archive, holding "
     "the documents its register entry lists. Blank where the application "
     "has no documents, or where they were acquired after the most recent "
     "sync."),
    ("Energy projects", "All columns",
     "Nationally significant infrastructure projects (energy) from the "
     "Planning Inspectorate register: metadata read from each project's "
     "page (snapshotted before parsing), no project documents fetched. "
     "'Stated capacity' is lifted verbatim from the page's description — "
     "the description column shows the sentence it came from. A blank "
     "capacity means the page states none, not that the project is small."),
    ("External aggregates", "All tables",
     "Aggregate figures published by Ofgem, NESO and UK Power Networks "
     "about data centre power demand, presented beside the planning-"
     "derived sheets and deliberately never joined to them: the sources "
     "measure different quantities (contracted grid headroom, developer "
     "survey responses, metered draw), none of which is the quantity a "
     "planning application states. Each external figure carries its "
     "source document, its table or paragraph number, and the date it "
     "was read; the verbatim quote is included where one was "
     "transcribed. Two of the sources are anonymised by their "
     "publishers, and no attempt has been made to match them to sites."),
    ("External aggregates", "Size distribution table",
     "Ofgem's Table 1 — the size distribution of the ~315 data centre "
     "projects holding ~73 GW of contracted connection offers in the GB "
     "demand queue at June 2025 — beside the count of sites in this "
     "workbook whose planning documents yield a figure in the same "
     "band. The two universes overlap but neither contains the other: "
     "the queue includes projects that have never filed a planning "
     "application, and this workbook includes the built estate back to "
     "2015. The workbook column is recomputed from the Sites sheet at "
     "every generation."),
    ("Capacity claims", "All columns",
     "One row per claim from a named external source — currently the "
     "119 transmission demand rows of NESO's Existing Agreements "
     "Register (published with the Gate 2 connections-reform results; "
     "snapshot and caveats in data/external_sources/README.md). Every "
     "figure is contracted grid connection capacity: a ceiling a "
     "developer once agreed with the grid operator, not IT load, not "
     "built capacity, and not what any site draws. The register is "
     "consent-based and records pre-reform positions, so absence "
     "proves nothing and entries can shrink or lapse. Most claims "
     "match no site in this workbook — that is the demand pipeline the "
     "planning system may not have seen yet, and the blank is the "
     "finding."),
    ("Capacity claims", "Matched site / confidence / method / evidence",
     "Where a claim names the same physical project as a site row, the "
     "attachment is a hand-adjudicated inference, not a join: 'strong' "
     "means name identity with no plausible competing referent; "
     "'probable' means independent details triangulate but the row name "
     "alone would not identify the site; 'tentative' means place and "
     "scale agree and nothing more — a lead, not an attribution. The "
     "evidence column is the written reasoning behind the match, and a "
     "register figure is never copied into the Sites sheet: where it "
     "diverges from the planning-derived figure for the same site, the "
     "divergence is the story."),
    ("Operator disclosure", "All columns",
     "One row per operator, counting the figures it has stated to each "
     "of five audiences: the planning authority (the site's own "
     "application documents), the grid operator (NESO's register), the "
     "auditors (accounts filed at Companies House), customers (the "
     "operator's own website) and the environmental regulator (the "
     "standby generator fleet in the site's environmental permit, a "
     "thermal rating in MWth rather than electrical megawatts). The "
     "counts are aggregates and the sheet "
     "names what each one covers: 'Which sites' lists the sites behind "
     "the count, and every figure counted here appears as its own row "
     "on Figures by audience with the source it was published in. "
     "Read it as a description of disclosure, not a scoreboard: almost "
     "none of these companies is obliged to publish capacity at all, "
     "and an empty column is not evidence of concealment."),
    ("Operator disclosure", "Audiences told / Terms the figures are "
     "published under",
     "'Audiences told' counts how many of the four have been given a "
     "figure. The terms column reproduces what each figure was "
     "published under, unconverted — 'Total Capacity', 'IT load' and "
     "'Total IT power' are not synonyms and are not merged. Where a "
     "figure was carried in a page's underlying data rather than in "
     "its text, the term is the data key it was carried under, which "
     "is why one operator's entry is a machine field name."),
    ("Figures by audience", "All columns",
     "One row per figure, for every site where more than one audience "
     "was given one. Figures are in the unit the source printed, and "
     "MVA is never converted to MW because no power factor is "
     "published. 'Source' opens the document or page the figure was "
     "published in, 'Where in the source' is the page or row within "
     "it, and 'What the source says' is the verbatim span the figure "
     "was read from — checked against the committed snapshot rather "
     "than transcribed. Planning-derived rows carry the project's own "
     "adjudication of the application documents and are marked "
     "'(planning-derived)' in the confidence column, because there is "
     "no external match to score."),
    ("Figures by audience", "Same quantity, different audience?",
     "Marks the narrow comparison where a gap really is a gap: one "
     "site, one quantity, more than one audience. Elsewhere a "
     "difference between figures is usually two different measurements "
     "— IT load is supposed to be smaller than total site power, and a "
     "contracted grid connection is a different thing again."),
]


def _git_commit() -> str:
    try:
        return subprocess.run(
            ["git", "rev-parse", "--short", "HEAD"],
            capture_output=True, text=True, cwd=Path(__file__).parent.parent,
        ).stdout.strip()
    except Exception:
        return "unknown"


def _hyperlink(url: str | None, label: str) -> str:
    """A clickable cell. Labels keep a wide sheet readable where a raw
    URL would not; the dictionary records each link column's target."""
    if not url:
        return ""
    return f'=HYPERLINK("{url}", "{label}")'


def _norm_key(s: str) -> str:
    """Alphanumerics only, lowercased. Folder names carry a
    filesystem-sanitised site key (slashes become spaces), so exact
    comparison fails on every key containing an application-style ref;
    stripping punctuation from both sides is what actually matches."""
    return re.sub(r"[^a-z0-9]+", "", s.lower())


def _drive_folder_map() -> dict[str, str]:
    """normalised site_key -> Drive folder URL, from the sync ledger.

    The ledger records every folder the sync created, keyed
    '<parent id>/<name>' with names shaped '<sanitised key> — <site name>'.
    Read-only here: a site missing from the ledger simply has not been
    synced yet, and its cell stays blank rather than guessing a URL.
    """
    if not DRIVE_LEDGER.exists():
        return {}
    try:
        folders = json.loads(DRIVE_LEDGER.read_text()).get("folders", {})
    except Exception:
        return {}
    roots = [v for k, v in folders.items() if k.endswith("/sites")]
    if not roots:
        return {}
    root = roots[0]
    out: dict[str, str] = {}
    for key, fid in folders.items():
        parent, _, name = key.partition("/")
        if parent != root:
            continue
        out[_norm_key(name.split(" — ")[0])] = (
            f"https://drive.google.com/drive/folders/{fid}")
    return out


def _drive_application_map() -> dict[tuple[str, str], str]:
    """(normalised site key, folder-name form of the ref) -> Drive URL.

    Application folders sit one level below their site's, named for the
    reference with slashes swapped for underscores (Drive treats a slash
    as a path separator). Linking straight to the application saves a
    reporter opening a site folder of forty siblings to find the one the
    row is about.

    Keyed on the site as well as the reference, because the same folder
    name legitimately appears under more than one site — a shared
    application between neighbouring councils, or two references that
    collide once truncated to the folder-name limit. Keying on the name
    alone silently sent some of those to another site's documents.

    As with the site map, absence means not-yet-synced, and the caller
    leaves the link out rather than constructing a URL that may 404.
    """
    if not DRIVE_LEDGER.exists():
        return {}
    try:
        folders = json.loads(DRIVE_LEDGER.read_text()).get("folders", {})
    except Exception:
        return {}
    roots = {v for k, v in folders.items() if k.endswith("/sites")}
    if not roots:
        return {}
    site_key_by_id = {
        fid: _norm_key(name.split(" — ")[0])
        for k, fid in folders.items()
        for parent, _, name in (k.partition("/"),) if parent in roots}
    out: dict[tuple[str, str], str] = {}
    for key, fid in folders.items():
        parent, _, name = key.partition("/")
        site_key = site_key_by_id.get(parent)
        if site_key:
            out[(site_key, name)] = f"https://drive.google.com/drive/folders/{fid}"
    return out


def _drive_application_url(app_map, site_key: str, ref: str) -> str:
    return app_map.get((_norm_key(site_key), clean_ref(ref)), "")


def _drive_findings_map() -> dict[str, str]:
    """normalised site_key -> Drive URL of that site's findings CSV.

    The folder maps above are built from the ledger's `folders`; this one
    from its `files`, which records an id per uploaded path. Sending a
    reporter to the folder and leaving them to spot one CSV among the
    application subfolders is a step that can be removed, and the panel
    that mentions the file is the natural place to remove it from.

    The site key comes from the containing folder rather than the
    filename, so it matches `_drive_folder_map` exactly even though both
    now carry it.

    Absent means not yet synced — a renamed CSV has no id until the sync
    has uploaded it, so a reader built between the rename and the sync
    simply describes the file without linking to it. That is the same
    contract as the folder maps and the reason the reader is rebuilt
    after a sync rather than before one.
    """
    if not DRIVE_LEDGER.exists():
        return {}
    try:
        files = json.loads(DRIVE_LEDGER.read_text()).get("files", {})
    except Exception:
        return {}
    out: dict[str, str] = {}
    for path, meta in files.items():
        name = PurePosixPath(path).name
        if not (name.startswith("_findings") and name.endswith(".csv")):
            continue
        fid = (meta or {}).get("id")
        if not fid:
            continue
        # A ledger entry whose local file has gone is a file the next
        # prune will bin — the old `_findings.csv` before the rename was
        # synced, for instance. Linking to it would hand a reporter a URL
        # that works today and 404s after the sync, which is worse than
        # no link. Existence locally is the cheapest proxy for "this is
        # the copy the current tree would upload".
        if not Path(path).exists():
            continue
        folder = PurePosixPath(path).parent.name
        out[_norm_key(folder.split(" — ")[0])] = (
            f"https://drive.google.com/file/d/{fid}/view")
    return out


def clean_ref(ref: str) -> str:
    """The staging tree's folder name for an application reference."""
    out = re.sub(r'[<>:"/\\|?*\x00-\x1f]+', " ", (ref or "").replace("/", "_"))
    return re.sub(r"\s+", " ", out).strip(" .")[:60].strip(" .")


def _haversine_km(lat1, lon1, lat2, lon2) -> float:
    p1, p2 = radians(lat1), radians(lat2)
    return 2 * 6371 * asin(sqrt(
        sin((p2 - p1) / 2) ** 2
        + cos(p1) * cos(p2) * sin(radians(lon2 - lon1) / 2) ** 2))


def main() -> None:
    # Nothing is written from adjudications nobody has
    # corrected. See dcp/adjudication_gate.py.
    adjudication_gate.require_corrected()
    ap = argparse.ArgumentParser()
    default_out = Path(
        f"data/exports/dc_build_handover_{dt.date.today().isoformat()}.xlsx")
    ap.add_argument("--out", type=Path, default=default_out)
    args = ap.parse_args()

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    from collections import defaultdict
    from urllib.parse import urlparse

    from dcp import capacity_claims as ccl
    from dcp import operator_disclosure as od
    from dcp import consumption_context as cc
    from dcp import external_aggregates as extagg
    from dcp import organisations
    from dcp import proposal
    from dcp import signals as sig
    from dcp import site_profile
    from dcp import site_scale as scale

    with db.connect() as conn, conn.cursor() as cur:
        cur.execute(SITE_SQL)
        site_rows = cur.fetchall()
        cur.execute(APP_SQL)
        app_rows = cur.fetchall()
        cur.execute(BARBOUR_ONLY_SQL)
        barbour_rows = cur.fetchall()
        cur.execute(NSIP_SQL)
        nsip_rows = cur.fetchall()

    # Environmental signals are extracted deterministically from the
    # description text (dcp/signals.py) rather than asked of the model:
    # reproducible, free, and carrying no risk to a validated prompt.
    # A floor, not a census — descriptions are terse, and the substantive
    # environmental content lives in the documents, which deep-read covers.
    app_env: dict[str, list[str]] = {}
    site_env: dict[str, set[str]] = defaultdict(set)
    site_desc: dict[str, list[str]] = defaultdict(list)
    for r in app_rows:
        site_key, ref, description = r[0], r[1], r[-1]
        found = sig.environmental_signals(description)
        flat = sig.flatten(found)
        app_env[ref] = flat
        site_env[site_key].update(found.keys())
        if description:
            site_desc[site_key].append(description)

    # Status & recency rollups, from the same rows the Applications sheet
    # shows so the two sheets cannot disagree. "Without recorded decision"
    # is exactly that: no decision date and no decided-type status text —
    # an indicator of liveness, defined verbatim in the dictionary.
    site_live: dict[str, int] = defaultdict(int)
    site_latest_app: dict[str, dt.date] = {}
    site_latest_decision: dict[str, dt.date] = {}
    site_portal: dict[str, tuple[dt.date | None, str]] = {}
    site_hosts: dict[str, set[str]] = defaultdict(set)
    site_refs: dict[str, list[str]] = defaultdict(list)
    for r in app_rows:
        site_key, ref = r[0], r[1]
        status, received, decided, url = r[4], r[5], r[6], r[12]
        site_refs[site_key].append(ref)
        if decided is None and not (status and _DECIDED_STATUS_RE.search(status)):
            site_live[site_key] += 1
        if received and (site_key not in site_latest_app
                         or received > site_latest_app[site_key]):
            site_latest_app[site_key] = received
        if decided and (site_key not in site_latest_decision
                        or decided > site_latest_decision[site_key]):
            site_latest_decision[site_key] = decided
        if url and not url.startswith("file://"):
            host = (urlparse(url).hostname or "").lower()
            if host:
                site_hosts[site_key].add(host)
            prior = site_portal.get(site_key)
            key_date = received or dt.date.min
            if prior is None or key_date >= (prior[0] or dt.date.min):
                site_portal[site_key] = (received, url)

    # Building floorspace per site, used only where no capacity has been
    # disclosed. Two deliberate restrictions, both learned the hard way:
    #
    #   - Only floorspace signal types. `site_area` and bare
    #     `development_scale` routinely carry land parcels, and one site
    #     came through at 117 km2 of "floor area" — run through a kW/m2
    #     factor that becomes a gigawatt estimate for a shed.
    #   - The median, not the max. A site's documents quote many areas
    #     (a phase, a hall, the whole scheme); the largest is the least
    #     representative, while the median tracks the building.
    # Derived signals shared with the web view (dcp/site_profile). Both
    # consumers call the same code so neither can present a different
    # answer for the same site.
    # Confirmed members only, so a proposal in the priors file changes
    # nothing about a build until a person has confirmed it.
    alias_index = organisations.alias_index(organisations.load_groups())
    with db.connect() as conn:
        site_profiles = site_profile.load_site_profiles(conn)
        coverage = site_profile.load_coverage(conn)
        # Prose counts drive the caveats; total counts stay for display.
        # The workbook and the reader must agree on which rows are
        # provisional, so both read this from site_profile rather than
        # deciding it locally.
        cov_detail = site_profile.load_coverage_detail(conn)

    # Shared with the reader, which used to pass None here and so showed
    # nothing for the 43 sites this figure covers. One loader, so the
    # two artefacts cannot disagree about a site's floor area.
    with db.connect() as conn:
        site_floorspace = scale.load_site_floorspace(conn)

    drive_urls = _drive_folder_map()

    # NSIP projects with usable coordinates, unpacked once for both the
    # per-site proximity columns and the Energy projects sheet.
    nsip = []
    for (ref, status, lat, lon, name, applicant, app_type, region, stage,
         caps, description, dev_site, accepted, url) in nsip_rows:
        cap_list = caps if isinstance(caps, list) else []
        nsip.append({
            "ref": ref, "status": status, "lat": lat, "lon": lon,
            "name": name or ref, "applicant": applicant or "",
            "app_type": app_type or "", "region": region or "",
            "stage": stage or "", "capacity": "; ".join(cap_list),
            "description": description or "", "developer_site": dev_site,
            "accepted": accepted or "", "url": url,
        })

    def nearest_energy(lat, lon):
        if lat is None or lon is None or not nsip:
            return None
        best = min(nsip, key=lambda p: _haversine_km(lat, lon, p["lat"], p["lon"]))
        return best, round(_haversine_km(lat, lon, best["lat"], best["lon"]), 1)

    wb = Workbook()

    def _sheet(title, headers):
        ws = wb.create_sheet(title) if wb.sheetnames != ["Sheet"] else wb.active
        ws.title = title
        ws.append(headers)
        for i in range(1, len(headers) + 1):
            c = ws.cell(row=1, column=i)
            c.font = Font(bold=True)
            c.fill = PatternFill("solid", fgColor="DDDDDD")
            c.alignment = Alignment(vertical="center")
        ws.freeze_panes = "A2"
        return ws

    # ---- Read me ---------------------------------------------------------
    ws = _sheet("Read me", ["Sheet", "Column", "What it contains and how it was derived"])
    for row in DICTIONARY:
        ws.append(list(row))
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 44
    ws.column_dimensions["C"].width = 120
    for r in ws.iter_rows(min_row=2):
        r[2].alignment = Alignment(wrap_text=True, vertical="top")

    # ---- Sites -----------------------------------------------------------
    ws = _sheet("Sites", SITE_HEADERS)
    site_coords: list[tuple[float, float, str]] = []
    # For the External aggregates sheet: the basis and value behind each
    # site row's headline figure, plus whether its prose was read in full.
    # Collected from the very estimates the Sites sheet displays, so the
    # aggregate table cannot disagree with the rows it summarises.
    agg_figures: list[tuple[str, float | None, bool]] = []
    # DESNZ consumption context: the series is loaded once and the
    # national change computed once, so every row's comparison is against
    # the same baseline. Coverage is asserted after the loops — every row
    # either maps or is counted, and the export prints both numbers.
    desnz = cc.load_series()
    desnz_national = round(cc.national_change(desnz))
    ctx_mapped = ctx_unmapped = 0
    ctx_unrecognised: set[str] = set()
    for r in site_rows:
        (key, cls, name, lat, lon, csrc, councils, n_apps, refs, verdicts,
         docs, findings_n, it_load_mw, total_site_mw, grid_mw, gen_mw,
         n_capacity, n_excluded, families, eia_ref, eia_doc, manual_docs,
         ptno, btitle, bstage, bvalue, bfloor, bsite, bplan, bdecision,
         bauthority) = r
        if lat is not None and lon is not None:
            site_coords.append((lat, lon, name or key))
        eia = " + ".join(
            label for hit, label in
            [(eia_ref, "ref pattern"), (eia_doc, "ES documents")] if hit)

        # Character from the site's own descriptions, rolled up by
        # significance; scale from the strongest evidence available, with
        # the basis stated so a floor-area inference is never mistaken for
        # a disclosed capacity.
        character = scale.rollup_character(
            [scale.character_for(d) for d in site_desc.get(key, ())]
            or [scale.character_for(name)])
        # One rankable figure with its qualifications; falls back through
        # disclosed IT load -> total site -> grid -> generation -> a
        # floorspace inference, losing authority at each step and saying so.
        prof = site_profiles.get(key, {})
        held, read = coverage.get(key, (docs or 0, 0))
        _cd = cov_detail.get(key, {})
        p_held = _cd.get("prose_held", held)
        p_read = _cd.get("prose_read", read)
        est = scale.power_estimate(
            it_load_mw=it_load_mw, total_site_mw=total_site_mw,
            grid_mw=grid_mw, generation_mw=gen_mw,
            floorspace_sqm=site_floorspace.get(key),
            has_documents=bool(docs),
            prose_held=p_held, prose_read=p_read)
        agg_figures.append((est.basis, est.value_mw,
                            p_held > 0 and p_read >= p_held))

        if est.value_mw is not None:
            band_key, band_label = scale.scale_from_mw(est.value_mw)
            basis = ("stated_capacity" if est.confidence in ("High", "Medium")
                     else "floor_area" if est.basis.startswith("Estimated")
                     else "stated_capacity")
        else:
            band_key, band_label, basis = "", "", "none"

        # A site with no member applications is a Barbour-anchored
        # pre-planning project: "no documents held" would be true but
        # misleading — no public material exists to hold.
        pre_app = (n_apps or 0) == 0
        cap_key, cap_label = site_profile.capacity_status(
            pre_application=pre_app, docs_held=p_held, docs_read=p_read,
            power_value_mw=est.value_mw, power_basis=est.basis)
        # "No capacity disclosed" from the estimator presumes the
        # documents were read. Where none have been (or none exist yet),
        # saying so in the power block too keeps a sort on Power basis
        # honest.
        power_basis_cell = est.basis
        power_caveat_cell = est.caveat
        # A figure from a partially-read site is a floor, not a measurement:
        # further reading can raise it but never lower it. Saying so on the
        # figure itself is the difference between a reader treating 500MW as
        # this site's capacity and treating it as the largest we have seen
        # so far.
        is_prov, prov_note = site_profile.provisional(p_held, p_read)
        # The register's own wording, reduced to the clause that describes
        # the development. Verbatim, so it stays quotable; the untouched
        # description remains on every Applications row.
        _summary, _descriptive = proposal.summarise(
            site_desc.get(key) or [btitle])
        proposal_cell = proposal.tidy(_summary)
        proposal_flag = ("Yes" if _descriptive else
                         "No — only procedural applications on record")
        if is_prov and est.value_mw is not None:
            power_basis_cell = f"{est.basis} {site_profile.PROVISIONAL_MARK}"
            power_caveat_cell = ((est.caveat + " ") if est.caveat else "") + prov_note
        if est.value_mw is None and cap_key in ("not_yet_analysed",
                                                "partially_analysed",
                                                "pre_application"):
            power_basis_cell = cap_label
            if cap_key != "pre_application":
                power_caveat_cell = ("Absence of a figure here is not a "
                                     "disclosure fact until analysis "
                                     "completes.")
        acq = site_profile.acquisition_status(
            pre_application=pre_app, docs_held=held,
            hosts=site_hosts.get(key, ()), refs=site_refs.get(key, ()))

        near = nearest_energy(lat, lon)
        near_name, near_km, near_cap = "", "", ""
        if near:
            p, km = near
            near_name = f"{p['name']} ({p['ref']})"
            near_km = km
            near_cap = p["capacity"]

        # Consumption context: emitted as a trio or not at all — an
        # authority name without its figures, or a national baseline
        # beside a blank, would invite the comparison the columns exist
        # to make without the terms that make it honest.
        ctx_la = cc.authority_for(councils, bauthority)
        ctx_pct = cc.change_pct(desnz[ctx_la]) if ctx_la else None
        if ctx_pct is not None:
            ctx_mapped += 1
            ctx_cells = [ctx_la, round(ctx_pct), desnz_national]
        else:
            ctx_unmapped += 1
            ctx_cells = ["", "", ""]
        ctx_unrecognised.update(cc.unrecognised(councils))

        portal = site_portal.get(key)
        n_hosts = len(site_hosts.get(key, ()))
        # "(1 of 5)" was true and unreadable: it never said what the five
        # were. A site can span councils, and each council has its own
        # register, so the link can only ever reach one of them — say
        # which thing is being counted.
        portal_label = ("Open portal" if n_hosts <= 1
                        else f"Open portal (1 of {n_hosts} registers)")

        row = [
            key, cls, name, proposal_cell, proposal_flag,
            lat, lon, csrc,
            ", ".join(councils or []),
            _hyperlink(drive_urls.get(_norm_key(key)), "Open Drive folder"),
            _hyperlink(portal[1] if portal else None, portal_label),
            n_apps, "\n".join(refs or []),
            site_live.get(key, 0),
            str(site_latest_app.get(key) or ""),
            str(site_latest_decision.get(key) or ""),
            ", ".join(sorted(verdicts or [])),
            est.value_mw, power_basis_cell, est.confidence, power_caveat_cell,
            ("Yes — reading incomplete" if is_prov else "No"),
            it_load_mw, total_site_mw, grid_mw, gen_mw,
            prof.get("gen_figure_basis") or "",
            prof.get("gen_unit_count") if prof.get("gen_unit_count") else "",
            prof.get("gen_unit_mw") if prof.get("gen_unit_mw") else "",
            n_capacity or "", n_excluded or "",
            scale.CHARACTERS[character].label, band_label,
            scale.BASIS_NOTE[basis],
            prof.get("generator_count") or "",
            prof.get("generator_fuel") or "",
            prof.get("generator_caveat") or "",
            prof.get("cooling_method") or "",
            prof.get("water_evidence") or "",
            prof.get("cooling_caveat") or "",
            prof.get("eia_status_label") or "", eia,
            ", ".join(sorted(site_env.get(key, ()))),
            prof.get("end_user") or "",
            prof.get("applicant_of_record") or "",
            prof.get("operator_group") or "",
            prof.get("advisers") or "",
            prof.get("named_in_documents") or "",
            prof.get("authority") or "",
            prof.get("parties_source") or "",
            # Already ordered by count in SQL; the tail is long and thin,
            # so show the families that actually characterise the site and
            # say how many more there are rather than filling the cell.
            (", ".join((families or [])[:TOP_FAMILIES])
             + (f"  (+{len(families) - TOP_FAMILIES} more)"
                if families and len(families) > TOP_FAMILIES else "")),
            held, read, findings_n, manual_docs or "",
            near_name, near_km, near_cap,
            *ctx_cells,
            cap_label, acq,
            ptno, btitle, bstage, bvalue, bfloor, bsite,
            str(bplan or ""), str(bdecision or ""),
        ]
        assert len(row) == len(SITE_HEADERS), \
            f"site row width {len(row)} != {len(SITE_HEADERS)}"
        ws.append(row)

    # Pre-planning rows the sites table does NOT already carry. The
    # blanks are the content — for these, manual determination is not a
    # fallback but the only route, which is exactly what the coverage
    # columns say.
    existing_keys = {r[0].upper() for r in site_rows}
    appended_barbour = 0
    barbour_only_parties: dict[str, tuple[str, dict]] = {}
    for (pref, title, pstage, dev_type, authority, address, description,
         plat, plon, pvalue, pfloor, psite, pplan, pdecision,
         praw) in barbour_rows:
        pseudo_key = f"PTNO-{pref}"
        if pseudo_key.upper() in existing_keys:
            continue
        appended_barbour += 1
        if plat is not None and plon is not None:
            site_coords.append((plat, plon, title or pseudo_key))
        cap_key, cap_label = site_profile.capacity_status(
            pre_application=True, docs_held=0, docs_read=0,
            power_value_mw=None, power_basis="")
        acq = site_profile.acquisition_status(
            pre_application=True, docs_held=0, hosts=(), refs=())
        near = nearest_energy(plat, plon)
        near_name, near_km, near_cap = "", "", ""
        if near:
            p, km = near
            near_name, near_km, near_cap = (
                f"{p['name']} ({p['ref']})", km, p["capacity"])
        env = sig.environmental_signals(description)
        bsummary, bdescriptive = proposal.summarise([description, title])
        bparties = site_profile.site_parties(
            site_profile.barbour_parties(praw or {}, str(pref or "")),
            (), [site_profile._AUTHORITY_PHONE_RE.sub("", authority or "")],
            alias_index)
        barbour_only_parties[pseudo_key] = (title or pseudo_key, bparties)
        ctx_la = cc.authority_for((), authority)
        ctx_pct = cc.change_pct(desnz[ctx_la]) if ctx_la else None
        if ctx_pct is not None:
            ctx_mapped += 1
            ctx_cells = [ctx_la, round(ctx_pct), desnz_national]
        else:
            ctx_unmapped += 1
            ctx_cells = ["", "", ""]
        row = [
            pseudo_key, "barbour_only", title,
            proposal.tidy(bsummary),
            "Yes" if bdescriptive else "No — Barbour intelligence only",
            plat, plon, "barbour", authority or "",
            _hyperlink(drive_urls.get(_norm_key(pseudo_key)), "Open Drive folder"),
            "",                       # no public register entry exists
            0, "", 0, "", "", "",
            None, cap_label, "", "", "n/a — no documents",
            None, None, None, None, "", "", "", "", "",
            "", "", "",               # character/scale unknowable pre-application
            "", "", "", "", "", "", "", "",
            ", ".join(sorted(env.keys())),
            bparties["end_user"], bparties["applicant_of_record"],
            bparties["operator_group"], bparties["advisers"],
            bparties["named_in_documents"], bparties["authority"],
            bparties["parties_source"],
            "", 0, 0, 0, "",
            near_name, near_km, near_cap,
            *ctx_cells,
            cap_label, acq,
            pref, title, pstage, pvalue, pfloor, psite,
            str(pplan or ""), str(pdecision or ""),
        ]
        assert len(row) == len(SITE_HEADERS), \
            f"barbour row width {len(row)} != {len(SITE_HEADERS)}"
        ws.append(row)

    # ---- Applications ------------------------------------------------------
    ws = _sheet("Applications", APP_HEADERS)
    drive_app_urls = _drive_application_map()
    for r in app_rows:
        vals = [str(x) if isinstance(x, (dt.date, dt.datetime)) else x for x in r]
        # A row that cannot reach its own documents sends the reader back
        # to the site folder to hunt among its siblings.
        folder = _hyperlink(
            _drive_application_url(drive_app_urls, r[0], r[1]) or None,
            "Open Drive folder")
        # Insert the derived signals column just before Address/Description.
        ws.append(vals[:13] + [folder] + vals[13:-2]
                  + ["\n".join(app_env.get(r[1], ()))] + vals[-2:])

    # ---- Energy projects ----------------------------------------------------
    # Ranked by distance to the nearest located data-centre site (main and
    # pre-planning rows alike): the co-location question is why the layer
    # exists. Distance is to the nearest *located* site — a large value
    # means far from the sites we can place, not provably remote.
    ws = _sheet("Energy projects", ENERGY_HEADERS)
    ranked = []
    for p in nsip:
        if site_coords:
            d, sname = min(
                ((_haversine_km(p["lat"], p["lon"], la, lo), nm)
                 for la, lo, nm in site_coords))
            ranked.append((round(d, 1), sname, p))
        else:
            ranked.append((None, "", p))
    ranked.sort(key=lambda t: (t[0] is None, t[0]))
    for d, sname, p in ranked:
        ws.append([
            p["ref"], p["name"], p["capacity"], p["app_type"], p["stage"],
            p["status"] or "", p["applicant"], p["region"], p["accepted"],
            sname, d,
            _hyperlink(p["url"], "PINS page"),
            _hyperlink(p["developer_site"], "Developer site"),
            p["description"][:500],
        ])

    # ---- External aggregates ------------------------------------------------
    # The premise is a domain fact established 2026-08-10 and recorded in
    # docs/EXTERNAL_DATA_SOURCES.md: none of these sources measures the
    # quantity a planning application states, so no external megawatt may
    # become a per-site column. Aggregates beside the data — never joined
    # to it — are the permitted form. External figures come from
    # dcp/external_aggregates (entered once, with provenance); everything
    # on this workbook's side of the comparison is computed from the
    # estimates the Sites sheet just displayed.
    ws = wb.create_sheet("External aggregates")

    def _agg_title(text):
        ws.append([text])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=12)

    def _agg_head(cols):
        ws.append(cols)
        for i in range(1, len(cols) + 1):
            c = ws.cell(row=ws.max_row, column=i)
            c.font = Font(bold=True)
            c.fill = PatternFill("solid", fgColor="DDDDDD")

    _agg_title("External aggregates: what the regulator and network "
               "operators publish about data centre power demand")
    ws.append(["These figures are presented beside the planning-derived "
               "sheets and deliberately never joined to them. The sources "
               "measure different quantities — the first table says what "
               "each one is and is not — and two of them are anonymised; "
               "no attempt has been made to match them to sites."])
    ws.append(["Every external figure names its source, its place in that "
               "source, and the date it was read. Every figure on this "
               "workbook's side is computed from the Sites sheet's own "
               "rows when the workbook is generated."])
    ws.append([])

    _agg_title("What each source of a megawatt figure measures")
    _agg_head(["Quantity", "Where it appears", "What it is",
               "What it is not"])
    for row in extagg.MEASURES:
        ws.append(list(row))
    ws.append([])

    _agg_title("Size distribution: the regulator's connection queue beside "
               "this workbook's planning documents")
    ws.append(["The two universes overlap but neither contains the other: "
               "the queue includes projects that have never filed a "
               "planning application — Ofgem's consultation argues much of "
               "it never will — while this workbook includes the built "
               "estate back to 2015 and sites below the queue's radar. "
               "The comparison shows what each side can see, not a "
               "shortfall to be subtracted."])
    _agg_head(["MW band", "Queue: data centre projects", "Queue: total MW",
               "Queue: share of queue MW",
               "This workbook: sites with a disclosed or plant-derived "
               "figure"])
    disclosed_mw = [v for b, v, _ in agg_figures
                    if b in scale.DISCLOSED_BASES and v is not None]
    for (label, _lo, _hi, n_proj, mw, pct), (_l, ours) in zip(
            extagg.OFGEM_QUEUE_BANDS, extagg.band_counts(disclosed_mw)):
        ws.append([label, n_proj, mw, pct, ours])
    ws.append(["All bands", extagg.OFGEM_QUEUE_TOTALS[0],
               extagg.OFGEM_QUEUE_TOTALS[1], "100%", len(disclosed_mw)])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
    ws.append(["Queue columns: Ofgem, Consultation Curate, Table 1 — "
               "contracted connection capacity at June 2025. Workbook "
               "column: each site's best-available basis (IT load, total "
               "site demand, grid connection or standby generation); "
               "floorspace estimates excluded because an inference does "
               "not belong in a column of contracted figures. The two "
               "columns deliberately have no MW total to compare: summing "
               "IT loads with generation capacities would manufacture a "
               "number no document states."])
    ws.append([])

    _agg_title("What the planning documents disclose "
               "(computed from this workbook's rows)")
    _agg_head(["What the documents yield", "Sites"])
    n_by_basis = {}
    for basis, _v, read_full in agg_figures:
        k = (basis, read_full) if basis == "No capacity disclosed" \
            else (basis, None)
        n_by_basis[k] = n_by_basis.get(k, 0) + 1
    disclosure_rows = [
        ("Headline figure is a disclosed IT load",
         n_by_basis.pop(("Disclosed IT load", None), 0)),
        ("Headline figure is a disclosed total site demand",
         n_by_basis.pop(("Disclosed total site demand", None), 0)),
        ("Headline figure is a grid connection capacity",
         n_by_basis.pop(("Grid connection capacity", None), 0)),
        ("Headline figure is inferred from standby generation plant",
         n_by_basis.pop(("Standby generation capacity", None), 0)),
        ("Headline figure is estimated from floorspace "
         "(an inference, not a disclosure)",
         n_by_basis.pop(("Estimated from floorspace", None), 0)),
        ("Readable documents read in full; no capacity or floorspace "
         "disclosed — for a data centre, itself notable",
         n_by_basis.pop(("No capacity disclosed", True), 0)),
        ("No capacity disclosed so far; reading incomplete, so the "
         "absence is provisional",
         n_by_basis.pop(("No capacity disclosed", False), 0)),
        ("No documents analysed yet",
         n_by_basis.pop(("Not yet analysed", None), 0)),
        ("No documents held",
         n_by_basis.pop(("No documents held", None), 0)),
    ]
    # A new basis string in site_scale must be classified here, not fall
    # off the sheet silently.
    assert not n_by_basis, f"bases missing from the aggregates sheet: " \
                           f"{sorted(n_by_basis)}"
    assert sum(n for _l, n in disclosure_rows) == len(agg_figures)
    for label, n in disclosure_rows:
        ws.append([label, n])
    ws.append(["Site rows summarised above", len(agg_figures)])
    ws.append(["Pre-planning rows appended from Barbour, not counted "
               "above (no public planning documents exist yet)",
               appended_barbour])
    ws.append([])

    _agg_title("Published aggregates")
    _agg_head(["What", "Figure", "Source", "Where in the source",
               "Verbatim quote"])
    for agg in extagg.AGGREGATES:
        src = extagg.SOURCES[agg.source_key]
        ws.append([agg.label, agg.figure,
                   f"{src.publisher} — {src.title}", agg.locator,
                   agg.quote])
    ws.append([])

    _agg_title("Sources")
    _agg_head(["Source", "Publisher", "Published", "Read on", "URL",
               "Notes"])
    for src in extagg.SOURCES.values():
        ws.append([src.title, src.publisher, src.published, src.accessed,
                   src.url, src.note])

    for col, width in (("A", 58), ("B", 30), ("C", 34), ("D", 34),
                       ("E", 64), ("F", 90)):
        ws.column_dimensions[col].width = width
    for row in ws.iter_rows(min_row=2):
        for c in row:
            if c.value is not None:
                c.alignment = Alignment(wrap_text=True, vertical="top")

    # ---- Parties -------------------------------------------------------
    # Long format, one row per organisation per role per site, because a
    # column per role would need sixty-three of them and would still
    # merge the two roles one firm holds on one scheme. The Sites sheet's
    # party columns are a summary of these rows; nothing appears there
    # that is not here.
    #
    # The count of organisations named exactly once is a row of its own
    # rather than a footnote: those names are excluded from both sheets
    # (dcp/site_profile.DOCUMENT_NAME_FLOOR), and a reader comparing a
    # site's parties against its findings CSV should be able to see how
    # many that was.
    ws = _sheet("Parties", [
        "Site key", "Site name", "Role", "Organisation",
        "Confirmed group", "Source", "Source ref", "Barbour role"])
    ROLE_LABELS = {
        "end_user": "End user", "applicant": "Applicant of record",
        "operator": "Operator", "adviser": "Adviser", "other": "Other role",
        "named_in_documents": "Named in documents"}
    n_party_rows = n_named_once = 0
    for r in site_rows:
        key, name = r[0], r[2]
        prof = site_profiles.get(key, {})
        n_named_once += prof.get("parties_named_once", 0)
        for party in prof.get("parties", ()):
            ws.append([key, name or key,
                       ROLE_LABELS.get(party.role, party.role),
                       party.name, party.group, party.source,
                       party.source_ref, party.barbour_role])
            n_party_rows += 1
        if prof.get("parties_named_once"):
            ws.append([key, name or key, "Named once, not listed", "",
                       "", "documents",
                       f"{prof['parties_named_once']} organisations", ""])
    for key, (name, bparties) in sorted(barbour_only_parties.items()):
        for party in bparties["parties"]:
            ws.append([key, name, ROLE_LABELS.get(party.role, party.role),
                       party.name, party.group, party.source,
                       party.source_ref, party.barbour_role])
            n_party_rows += 1
    for col, width in (("A", 34), ("B", 44), ("C", 20), ("D", 44),
                       ("E", 26), ("F", 12), ("G", 18), ("H", 30)):
        ws.column_dimensions[col].width = width
    print(f"  Parties: {n_party_rows} rows, "
          f"{n_named_once} single-mention names not listed")

    # ---- Capacity claims ----------------------------------------------------
    # Site-level external figures, the other permitted form beside the
    # aggregates above: each row is a claim as its source states it, and
    # where a hand-adjudicated match attaches it to a site the match's
    # confidence, method and written evidence are columns — the match is
    # our inference, so its reasoning ships with it. Deliberately a
    # separate sheet: the Sites sheet's power columns stay
    # planning-derived only, and the divergence between a register figure
    # and a planning figure for the same site is a finding, not an error.
    ws = _sheet("Capacity claims", [
        "Register entry", "MW", "Quantity", "Connection point",
        "Connection date", "Register as at", "Register row",
        "Matched site", "Match confidence", "Match method",
        "Match evidence", "Source", "Source URL"])
    with db.connect() as conn, conn.cursor() as cur:
        claim_rows = ccl.load_claim_rows(cur)
    for c in claim_rows:
        ws.append([
            c["claim_name"], c["value_mw"],
            ccl.QUANTITY_LABELS.get(c["quantity_type"], c["quantity_type"]),
            c["connection_point"], c["connection_date"],
            c["as_at"].isoformat() if c["as_at"] else None,
            c["source_locator"],
            c["site_name"] or c["site_key"],
            c["confidence"], c["method"], c["evidence"],
            ccl.SOURCE_TITLES.get(c["source_key"], c["source_key"]),
            c["source_url"]])
    for col, width in (("A", 40), ("B", 9), ("C", 24), ("D", 40), ("E", 14),
                       ("F", 14), ("G", 12), ("H", 44), ("I", 14), ("J", 22),
                       ("K", 90), ("L", 34), ("M", 48)):
        ws.column_dimensions[col].width = width
    for row in ws.iter_rows(min_row=2):
        for c in row:
            if c.value is not None:
                c.alignment = Alignment(wrap_text=True, vertical="top")
    n_claim_matches = sum(1 for c in claim_rows if c["confidence"])
    print(f"  Capacity claims: {len(claim_rows)} claims, "
          f"{n_claim_matches} matched to sites")

    # ---- Operator disclosure ------------------------------------------------
    # The same companies, across every audience the store holds. Computed
    # from the claims so it cannot disagree with the Capacity claims
    # sheet or the site panels; see dcp/operator_disclosure.py for why
    # the planning column is read from power_adjudication rather than
    # from claims. The columns come from od.AUDIENCES rather than being
    # written out here, so adding a source adds a column instead of
    # quietly dropping one.
    _aud_cols = [f"Told {lbl[0].lower()}{lbl[1:]}" for _k, lbl, _d
                 in od.AUDIENCES]
    ws = _sheet("Operator disclosure", [
        "Operator", "Audiences told", "Sites in this dataset",
        "Which sites", *_aud_cols,
        "Terms the figures are published under"])
    with db.connect() as conn, conn.cursor() as cur:
        op_rows = od.load_rows(cur)
        op_divs = od.load_divergences(cur)

    def _cell(row, key):
        got = row.by_audience.get(key) or []
        if not got:
            return "—"
        return f"{len(got)} figure{'' if len(got) == 1 else 's'}"

    for r in op_rows:
        # The count and the list, side by side: a bare "6 sites" is an
        # assertion, and the Figures by audience sheet is where each of
        # those sites' figures is itemised with its source.
        ws.append([r.operator, r.audiences, len(r.sites),
                   "; ".join(n for _k, n in r.site_names) or "—",
                   *[_cell(r, k) for k, _l, _d in od.AUDIENCES],
                   "; ".join(t for t in sorted(r.terms) if t) or "—"])
    ws.append([])
    ws.append([od.FAIRNESS_NOTE])
    ws.append([od.METHOD_NOTE])
    _widths = [30, 14, 20, 60] + [24] * len(od.AUDIENCES) + [60]
    for i, width in enumerate(_widths):
        ws.column_dimensions[get_column_letter(i + 1)].width = width
    for row in ws.iter_rows(min_row=2):
        for c in row:
            if c.value is not None:
                c.alignment = Alignment(wrap_text=True, vertical="top")

    # ---- Figures by audience ------------------------------------------------
    ws = _sheet("Figures by audience", [
        "Site", "Audience", "Figure", "Unit", "Quantity",
        "Published as", "Named in the source as", "Match confidence",
        "Same quantity, different audience?",
        "Source", "Where in the source", "What the source says"])
    for d in op_divs:
        lfl = {q["quantity_type"] for q in d.get("like_for_like", [])}
        for c in d["claims"]:
            ws.append([
                d["site"], dict(
                    (k, lbl) for k, lbl, _ in od.AUDIENCES).get(
                        c["audience"], c["audience"]),
                c["value"], c["unit"], c["quantity_type"],
                c["term"] or "—", c["claim_name"],
                c["confidence"] or "(planning-derived)",
                "yes" if c["quantity_type"] in lfl else "",
                _hyperlink(c.get("source_url"), "Open the source"),
                c.get("locator") or "—",
                c.get("quote") or "—"])
    for col, width in (("A", 46), ("B", 24), ("C", 10), ("D", 8), ("E", 20),
                       ("F", 26), ("G", 40), ("H", 18), ("I", 16),
                       ("J", 18), ("K", 20), ("L", 70)):
        ws.column_dimensions[col].width = width
    print(f"  Operator disclosure: {len(op_rows)} operators, "
          f"{len(op_divs)} sites told more than one audience")

    # ---- Provenance --------------------------------------------------------
    ws = _sheet("Provenance", ["Field", "Value"])
    with db.connect() as conn, conn.cursor() as cur:
        cur.execute("SELECT count(*) FROM documents")
        n_docs = cur.fetchone()[0]
        cur.execute("""SELECT count(DISTINCT document_id) FROM deepread_log
                       WHERE read_state = 'read'""")
        n_read = cur.fetchone()[0]
        cur.execute("""SELECT count(DISTINCT document_id) FROM deepread_log
                       WHERE read_state = 'read' AND model LIKE 'mlx%%'""")
        n_second = cur.fetchone()[0]
    pct = (100 * n_read // n_docs) if n_docs else 0
    for k, v in [
        ("RELEASE", "Phase 1 of 3"),
        ("What Phase 1 contains",
         f"Every site, application and energy project we hold, with all "
         f"analysis complete against the documents read so far. "
         f"{n_read:,} of {n_docs:,} documents ({pct}%) have been analysed."),
        ("What is NOT in Phase 1",
         f"{n_docs - n_read:,} documents are held but not yet analysed, so "
         "findings-derived columns (power, cooling, EIA status, parties, "
         "finding subjects) are silent for them. Per-site, 'Capacity status' "
         "and 'Documents analysed' say exactly which rows this affects — an "
         "empty power figure on an unanalysed site is not a disclosure fact. "
         "A small tail of applications is also still being retrieved."),
        ("Phase 2 (expected within days)",
         "The remaining documents analysed, and the acquisition tail "
         "completed. Every figure here is regenerated; nothing is "
         "overwritten, so Phase 1 stays auditable."),
        ("Phase 3", "A second, independent model's reading of the whole "
         f"corpus for corroboration ({n_second:,} documents so far)."),
        ("Water figures", "Deliberately not published as volumes. The water "
         "and cooling findings are dominated by drainage and flood "
         "engineering; only 93 sites disclose anything about consumption. "
         "'Cooling method' is reported instead, being both better evidenced "
         "and the thing that determines water demand."),
        ("Generated at (UTC)", dt.datetime.now(dt.timezone.utc)
                                 .isoformat(timespec="seconds")),
        ("Pipeline commit", _git_commit()),
        ("Sites (active)", len(site_rows)),
        ("Pre-planning projects (Barbour, no application)",
         f"{len(barbour_rows)} ({len(barbour_rows) - appended_barbour} materialised as sites, "
         f"{appended_barbour} appended by the exporter"),
        ("Applications in sites", len(app_rows)),
        ("Energy projects (NSIP layer)", len(nsip)),
        ("Documents in corpus", n_docs),
        ("Documents analysed by deep-read", n_read),
        ("Verdict column source", "v1 triage rubric — dc_build v2.1 "
         "catalogue sweep pending; column will switch to dc_build classes"),
        ("Barbour columns", "Barbour ABI, licensed, credit required in "
         "published output; contact/role fields deliberately excluded"),
        ("Findings columns", "verified deep-read findings (v1, quote-gated); "
         "coverage partial until v2 deep-read"),
        ("Known retrieval gaps", "Coventry (planandregulatory.coventry.gov.uk): "
         "28 applications behind bot protection, documents not retrievable "
         "by this pipeline and deliberately not worked around. Wiltshire "
         "PL/2022/09577: register entry requires a consultee login."),
        ("Energy projects scope", "Planning Inspectorate project-page "
         "metadata only (pages snapshotted before parsing); DCO document "
         "sets deliberately not fetched. A blank stated capacity means "
         "the page states none."),
    ]:
        ws.append([k, str(v)])

    args.out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(args.out)
    # No silent gaps: every Sites row either carries the DESNZ context or
    # is counted here, and a council prefix the mapping has never seen is
    # named rather than folded into the unmapped number.
    assert ctx_mapped + ctx_unmapped == len(site_rows) + appended_barbour, \
        (ctx_mapped, ctx_unmapped, len(site_rows), appended_barbour)
    print(f"Wrote {args.out}")
    print(f"  Sites: {len(site_rows)} + {appended_barbour} pre-planning; "
          f"Applications: {len(app_rows)}; Energy projects: {len(nsip)}")
    print(f"  Consumption context: {ctx_mapped} rows mapped to a DESNZ "
          f"authority, {ctx_unmapped} unmapped (spans several authorities, "
          f"Northern Ireland, development corporations, or no authority "
          f"recorded)")
    if ctx_unrecognised:
        print(f"  Consumption context: UNRECOGNISED council prefixes — add "
              f"to dcp/consumption_context.py: {sorted(ctx_unrecognised)}")


if __name__ == "__main__":
    main()
