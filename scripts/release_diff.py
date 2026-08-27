#!/usr/bin/env python3
"""Diff a build against the last release, by counting what a reader can reach.

A release is checked by comparing it with the one before, and until now
that comparison was done by eye. It found four regressions in 2.2 —
card links that did nothing, a chip that squashed the map, a dead energy
checkbox, two definitions of "prose" on one page — and it would have
found the 43 floorspace sites that vanished from the reader while the
workbook kept them, had anyone counted the basis column. An eye gets
tired; this does not.

It compares the three artefacts a release folder holds, whichever of
them both sides have:

  reader.html       tabs and views; site rows by key; links reachable
                    from each site's panel; table rows per view; section
                    and box headings; the filter controls a reporter can
                    pick; the header stamp's own numbers
  *.xlsx            sheets; rows and columns per sheet; column headings
                    added and removed
  *.duckdb          tables; rows per table

and, if the two priors files the redesign introduces exist, that every
`site_key` they name resolves to a site row in the build — a site that
was split or retired since the evidence was written is otherwise a
silent dangling reference.

The rule is asymmetric on purpose. Anything that **fell** — a tab gone,
a column gone, a site that lost links, fewer rows — exits non-zero,
because a build is supposed to add. Growth is reported and allowed.
A deliberate removal is declared with `--allow-fewer`, which prints the
same table and exits zero; the print is the record that someone meant
it.

    scripts/release_diff.py data/exports/phase2.3_build
    scripts/release_diff.py /tmp/reader.html --against data/exports/phase2.2_build
    scripts/release_diff.py BUILD --against RELEASE --allow-fewer

Counts, not content: it says a site panel has fewer links than it had,
not which one went. That is what the smoke test and a human are for.
"""

from __future__ import annotations

import argparse
import html
import re
import sys
from dataclasses import dataclass, field
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent))

from dcp import release

PRIORS_WITH_SITE_KEYS = (
    Path("data/priors/cohort_checks.yaml"),
    Path("data/priors/organisation_aliases.yaml"),
)


# --- the reader --------------------------------------------------------

@dataclass
class ReaderShape:
    tabs: list[str] = field(default_factory=list)
    views: list[str] = field(default_factory=list)
    site_keys: set[str] = field(default_factory=set)
    links_per_site: dict[str, int] = field(default_factory=dict)
    rows_per_view: dict[str, int] = field(default_factory=dict)
    sections: int = 0
    boxes: int = 0
    controls: list[str] = field(default_factory=list)
    stamp: dict[str, int] = field(default_factory=dict)
    dictionary_entries: int = 0


_TAB_RE = re.compile(r'<button id="tab-([a-z]+)"')
_VIEW_RE = re.compile(r'<section id="view-([a-z]+)"(.*?)</section>', re.DOTALL)
_SITE_RE = re.compile(r'<tr class="site" data-key="([^"]+)"')
# A site's panel is the <tr class="detail"> that follows its row. Links
# are counted there, not in the row, because the row's links are the
# same four on every site and the panel is where a reporter follows the
# evidence outward.
# The detail cell holds an applications TABLE, so `(.*?)</tr>` stops at
# that table's first row and counts only the links above it. It read as
# a third of every panel's links vanishing when the site page moved the
# applications into the left column (2026-08-25) — 2,287 to 1,526, while
# the panels had in fact gained links, 11,469 to 12,497. A panel now runs
# to the next site row, which is the only boundary that nests nothing.
_SITE_ROW_RE = re.compile(r'<tr class="site" data-key="([^"]+)"')
_DETAIL_START_RE = re.compile(r'<tr class="detail">')
_LINK_RE = re.compile(r'<a\s[^>]*href=')
_ROW_RE = re.compile(r'<tr[\s>]')
# Attributes, not a literal tag. Adding id="package" to one heading
# made this count fall by one and reported a section as lost when
# nothing had moved — a detector that a formatting change can trip
# spends the reviewer's attention on itself (2026-08-25).
_SECTION_RE = re.compile(r'<h2\b[^>]*\bclass="sec"')
_BOX_RE = re.compile(r'<h4>')
_OPTION_RE = re.compile(r'<option[^>]*>([^<]*)')
_CHECKBOX_RE = re.compile(r'<label[^>]*><input type="checkbox"[^>]*>\s*([^<]+)')
_STAMP_RE = re.compile(r'<div class="sub">(.*?)</div>', re.DOTALL)
_STAMP_NUM_RE = re.compile(r'([\d,]+)\s+([a-z][a-z ]+?)(?:\s*·|\s*$)')
_DICT_ENTRY_RE = re.compile(r'<div class="entry"')


_FILTERBAR_RE = re.compile(
    r'<div id="filterbar"[^>]*>(.*?)\n</div>', re.S)


def reader_shape(path: Path) -> ReaderShape:
    src = path.read_text(encoding="utf-8")
    shape = ReaderShape()
    shape.tabs = _TAB_RE.findall(src)
    views = _VIEW_RE.findall(src)
    shape.views = [v for v, _ in views]
    # The filter bar moved out of the sites view and above both the table
    # and the map, which they now share (2026-08-25). Scraped from
    # wherever it is rather than from inside a view: reading it out of
    # #view-sites reported every control as removed, and would have gone
    # on reporting nothing at all if one really were.
    bar = _FILTERBAR_RE.search(src)
    if bar:
        body = bar.group(1)
        shape.controls = ([html.unescape(o).strip() for o in _OPTION_RE.findall(body)]
                          + [html.unescape(c).strip() for c in _CHECKBOX_RE.findall(body)])
    for view, body in views:
        shape.rows_per_view[view] = len(_ROW_RE.findall(body))
        if view == "sites" and not bar:
            shape.controls = ([html.unescape(o).strip() for o in _OPTION_RE.findall(body)]
                              + [html.unescape(c).strip() for c in _CHECKBOX_RE.findall(body)])
        if view == "dict":
            shape.dictionary_entries = len(_DICT_ENTRY_RE.findall(body))
    shape.site_keys = set(_SITE_RE.findall(src))
    rows = [(m.group(1), m.end()) for m in _SITE_ROW_RE.finditer(src)]
    for i, (key, pos) in enumerate(rows):
        d = _DETAIL_START_RE.search(src, pos)
        if not d:
            continue
        if i + 1 < len(rows):
            end = rows[i + 1][1]
        else:
            # The LAST panel must stop at its table's end, not the end
            # of the document: run to len(src) and the slice swallows
            # every later view's links, so the final site's count moves
            # whenever anything after the table does. On 2026-08-27 it
            # attributed a -154 from the post-table views to the NPL
            # panel, whose own links had in fact gone from 3 to 4.
            t = src.find("</tbody>", d.start())
            end = t if t != -1 else len(src)
        shape.links_per_site[key] = len(_LINK_RE.findall(src[d.start():end]))
    shape.sections = len(_SECTION_RE.findall(src))
    shape.boxes = len(_BOX_RE.findall(src))
    m = _STAMP_RE.search(src)
    if m:
        text = html.unescape(re.sub(r"\s+", " ", m.group(1)))
        for num, label in _STAMP_NUM_RE.findall(text):
            shape.stamp[label.strip()] = int(num.replace(",", ""))
    return shape


# --- the workbook and the database ---------------------------------------

def workbook_shape(path: Path) -> dict[str, tuple[int, list[str]]]:
    """Sheet -> (data rows, column headings)."""
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True)
    out = {}
    for ws in wb.worksheets:
        headings = [str(c.value) for c in next(ws.iter_rows(min_row=1, max_row=1))
                    if c.value is not None]
        out[ws.title] = (max(ws.max_row - 1, 0), headings)
    return out


def duckdb_shape(path: Path) -> dict[str, int]:
    import duckdb
    con = duckdb.connect(str(path), read_only=True)
    tables = [t for (t,) in con.execute(
        "SELECT table_name FROM information_schema.tables "
        "WHERE table_schema='main' ORDER BY 1").fetchall()]
    return {t: con.execute(f'SELECT count(*) FROM "{t}"').fetchone()[0] for t in tables}


# --- comparison -----------------------------------------------------------

@dataclass
class Report:
    lines: list[str] = field(default_factory=list)
    fell: list[str] = field(default_factory=list)

    def row(self, what: str, before, after, *, fell_if_less: bool = True) -> None:
        mark = ""
        if isinstance(before, int) and isinstance(after, int):
            if after < before:
                mark = "  FELL"
                if fell_if_less:
                    self.fell.append(f"{what}: {before} -> {after}")
            elif after > before:
                mark = "  +"
        self.lines.append(f"  {what:<46} {before!s:>10} -> {after!s:<10}{mark}")

    def missing(self, what: str, items) -> None:
        items = sorted(items)
        if items:
            self.fell.append(f"{what}: {', '.join(items)}")
            self.lines.append(f"  {what}: REMOVED {', '.join(items)}")

    def added(self, what: str, items) -> None:
        items = sorted(items)
        if items:
            self.lines.append(f"  {what}: added {', '.join(items)}")


def compare_readers(rep: Report, before: ReaderShape, after: ReaderShape) -> None:
    rep.lines.append("reader.html")
    rep.missing("tabs", set(before.tabs) - set(after.tabs))
    rep.added("tabs", set(after.tabs) - set(before.tabs))
    rep.missing("views", set(before.views) - set(after.views))
    rep.added("views", set(after.views) - set(before.views))
    for view in sorted(set(before.rows_per_view) | set(after.rows_per_view)):
        rep.row(f"rows in view '{view}'", before.rows_per_view.get(view, 0),
                after.rows_per_view.get(view, 0))
    rep.row("site rows", len(before.site_keys), len(after.site_keys))
    gone = before.site_keys - after.site_keys
    # Sites legitimately leave when partitions or retirements move them;
    # that is still a removal and is declared with --allow-fewer.
    rep.missing("site keys", list(gone)[:12] + (["…"] if len(gone) > 12 else []))
    rep.added("site keys", [f"{len(after.site_keys - before.site_keys)} new"]
              if after.site_keys - before.site_keys else [])
    common = before.site_keys & after.site_keys
    lost = [k for k in common
            if after.links_per_site.get(k, 0) < before.links_per_site.get(k, 0)]
    # Counted as "kept", not "lost", so the number FALLS when panels lose
    # links. Written the other way round it read `0 -> 15`, which rep.row
    # scores as a rise and marks `+` — so on 2026-08-26 fifteen panels
    # each lost a link, the report said "nothing fell", and the exit code
    # was 0. The one check written to catch this could not catch it,
    # because the metric was phrased so that the bad direction was up.
    rep.row("site panels keeping every link they had", len(common),
            len(common) - len(lost))
    if lost:
        rep.lines.append("    lost a link: " + ", ".join(
            f"{k} {before.links_per_site[k]}->{after.links_per_site[k]}" for k in lost[:6])
            + (f" … and {len(lost) - 6} more" if len(lost) > 6 else ""))
    rep.row("links across all site panels",
            sum(before.links_per_site.values()), sum(after.links_per_site.values()))
    rep.row("section headings (h2.sec)", before.sections, after.sections)
    rep.row("box headings (h4)", before.boxes, after.boxes)
    rep.row("data dictionary entries", before.dictionary_entries, after.dictionary_entries)
    rep.missing("filter controls", set(before.controls) - set(after.controls))
    rep.added("filter controls", set(after.controls) - set(before.controls))
    for label in sorted(set(before.stamp) | set(after.stamp)):
        # The stamp's own numbers move with the corpus and are reported,
        # never judged: fewer documents held is a corpus fact, not a
        # build regression.
        rep.row(f"stamp: {label}", before.stamp.get(label, 0), after.stamp.get(label, 0),
                fell_if_less=False)


def compare_workbooks(rep: Report, before: dict, after: dict) -> None:
    rep.lines.append("workbook")
    rep.missing("sheets", set(before) - set(after))
    rep.added("sheets", set(after) - set(before))
    for sheet in sorted(set(before) & set(after)):
        b_rows, b_cols = before[sheet]
        a_rows, a_cols = after[sheet]
        rep.row(f"'{sheet}' rows", b_rows, a_rows)
        rep.row(f"'{sheet}' columns", len(b_cols), len(a_cols))
        rep.missing(f"'{sheet}' columns", set(b_cols) - set(a_cols))
        rep.added(f"'{sheet}' columns", set(a_cols) - set(b_cols))


def compare_duckdbs(rep: Report, before: dict, after: dict) -> None:
    rep.lines.append("duckdb")
    rep.missing("tables", set(before) - set(after))
    rep.added("tables", set(after) - set(before))
    for table in sorted(set(before) & set(after)):
        rep.row(f"'{table}' rows", before[table], after[table])


def check_priors(rep: Report, site_keys: set[str]) -> None:
    """Every site_key a priors file names must still be a site."""
    import yaml
    for path in PRIORS_WITH_SITE_KEYS:
        if not path.exists():
            continue
        doc = yaml.safe_load(path.read_text()) or {}
        named = set(_site_keys_in(doc))
        dangling = named - site_keys
        rep.lines.append(f"{path}")
        rep.row("site keys named", len(named), len(named), fell_if_less=False)
        rep.missing("site keys that no longer resolve", dangling)


def _site_keys_in(node):
    """Every value under a `site_key` key, at any depth."""
    if isinstance(node, dict):
        for k, v in node.items():
            if k == "site_key" and isinstance(v, str):
                yield v
            else:
                yield from _site_keys_in(v)
    elif isinstance(node, list):
        for item in node:
            yield from _site_keys_in(item)


# --- locating artefacts ---------------------------------------------------

def _find(folder_or_file: Path, kind: str) -> Path | None:
    if folder_or_file.is_file():
        return folder_or_file if folder_or_file.suffix == kind or (
            kind == ".html" and folder_or_file.suffix == ".html") else None
    matches = sorted(folder_or_file.glob(f"*{kind}"))
    return matches[-1] if matches else None


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("build", type=Path,
                    help="a release folder, or a single reader.html / .xlsx / .duckdb")
    ap.add_argument("--against", type=Path, default=None,
                    help="the release to compare with; defaults to the newest "
                         "data/exports/*_build folder other than the build itself")
    ap.add_argument("--allow-fewer", action="store_true",
                    help="declare that something was removed on purpose: print the "
                         "same table, exit 0")
    args = ap.parse_args()

    against = args.against
    if against is None:
        candidates = [d for d in release.release_dirs() if d.resolve() != args.build.resolve()]
        if not candidates:
            print("no release folder to compare against; pass --against", file=sys.stderr)
            return 2
        against = candidates[0]
    print(f"build:   {args.build}\nagainst: {against}\n")

    rep = Report()
    compared = 0
    for kind, shaper, comparer in ((".html", reader_shape, compare_readers),
                                   (".xlsx", workbook_shape, compare_workbooks),
                                   (".duckdb", duckdb_shape, compare_duckdbs)):
        b, a = _find(against, kind), _find(args.build, kind)
        if b is None or a is None:
            rep.lines.append(f"{kind[1:]}: not on both sides, skipped")
            continue
        try:
            before, after = shaper(b), shaper(a)
        except Exception as e:  # noqa: BLE001 — any unreadable artefact is a report line, not a crash
            rep.lines.append(f"{kind[1:]}: could not read ({str(e).splitlines()[0][:120]}); skipped")
            continue
        comparer(rep, before, after)
        compared += 1
        if kind == ".html":
            check_priors(rep, after.site_keys)

    print("\n".join(rep.lines))
    if compared == 0:
        print("\nnothing compared", file=sys.stderr)
        return 2
    if rep.fell:
        print("\nFELL:")
        for f in rep.fell:
            print(f"  {f}")
        if args.allow_fewer:
            print("\n--allow-fewer given: removals declared deliberate")
            return 0
        return 1
    print("\nnothing fell")
    return 0


if __name__ == "__main__":
    sys.exit(main())
