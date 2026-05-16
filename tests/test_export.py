"""Light smoke tests for the reporter export. Cover the two surfaces most
likely to break silently: the markdown header text formatting (`.format()`
on `METHODOLOGY_NOTE` is a sharp edge — literal braces have to be escaped),
and the xlsx structure (headers, row count, freeze pane, auto-filter)."""

from __future__ import annotations

import datetime as dt

import pytest

from dcp import export, worklist


@pytest.fixture
def fake_data() -> worklist.WorklistData:
    summary = {
        "total": 100, "dc": 30, "adjacent": 10,
        "unrelated": 55, "unknown": 5, "worklist": 40,
    }
    rows = [
        {
            "id": 1, "application_ref": "TestCouncil/24/0001",
            "description": "Outline planning application for a hyperscale data centre.",
            "address": "Site at Test Lane, Testtown TS1 1AA",
            "date_received": dt.date(2024, 4, 15),
            "url": "https://example.invalid/app/0001",
            "raw_metadata": {"area_name": "TestCouncil", "app_type": "Outline"},
            "council_name": "Test Council",
            "verdict": "DC", "worth_deep_read": "yes", "confidence": "sure",
            "signals": ["energy centre", "gas turbine", "BESS"],
            "why": "Explicit DC build with on-site generation language.",
            "discovered_via": ["dc_keyword", "foxglove_top10"],
            "foxglove": True,
            "tier1_hits": 2, "storage_hits": 1, "backup_hits": 0,
        },
        {
            "id": 2, "application_ref": "TestCouncil/24/0002",
            "description": "Variation of conditions",
            "address": "Other Site",
            "date_received": dt.date(2024, 1, 1),
            "url": None,
            "raw_metadata": {"area_name": "TestCouncil", "app_type": "Conditions"},
            "council_name": "Test Council",
            "verdict": "adjacent", "worth_deep_read": "maybe", "confidence": "guessing",
            "signals": ["substation"],
            "why": "Sparse description.",
            "discovered_via": ["spatial:TestCouncil/24/0001"],
            "foxglove": False,
            "tier1_hits": 0, "storage_hits": 0, "backup_hits": 0,
        },
    ]
    anchors = {
        "TestCouncil/24/0001": {
            "description": "Outline planning application for a hyperscale data centre.",
            "address": "Site at Test Lane, Testtown TS1 1AA",
        }
    }
    return worklist.WorklistData(
        summary=summary, rows=rows, anchors=anchors, model="fake-granite",
    )


def test_render_markdown_escapes_literal_braces_in_methodology(fake_data):
    """Regression guard: `{DC, adjacent}` and `{yes, maybe}` are literal
    set notation in the methodology paragraph; if `.format()` tries to
    interpret them as substitution keys we get KeyError at runtime."""
    generated_at = dt.datetime(2026, 5, 15, 12, 0, 0)
    md = export.render_markdown(
        data=fake_data, top=10, model="fake-granite", generated_at=generated_at,
    )
    assert "{DC, adjacent}" in md
    assert "{yes, maybe}" in md
    # And the substitution that IS meant to fire was applied:
    assert "the top 2 curated" in md or "top 2 curated" in md
    # The card was included
    assert "TestCouncil/24/0001" in md
    assert "Foxglove top-10" in md


def test_render_markdown_top_caps_card_count(fake_data):
    """`top=1` includes only the first ranked card; the universe summary
    still reports the full worklist size."""
    generated_at = dt.datetime(2026, 5, 15, 12, 0, 0)
    md = export.render_markdown(
        data=fake_data, top=1, model="fake-granite", generated_at=generated_at,
    )
    assert "TestCouncil/24/0001" in md
    assert "TestCouncil/24/0002" not in md
    assert "Worklist size:** 40" in md


def test_write_xlsx_structure(tmp_path, fake_data):
    """The xlsx has the expected headers, row count, freeze pane, and
    auto-filter set across all data rows."""
    from openpyxl import load_workbook

    path = tmp_path / "worklist.xlsx"
    export.write_xlsx(
        path=path, data=fake_data, model="fake-granite",
        generated_at=dt.datetime(2026, 5, 15, 12, 0, 0),
    )
    wb = load_workbook(path)
    assert "Worklist" in wb.sheetnames
    assert "Methodology" in wb.sheetnames
    ws = wb["Worklist"]
    assert [c.value for c in ws[1]][:5] == [
        "Rank", "Application ref", "Highlight",
        "Primary cohort", "Also in cohorts",
    ]
    # 2 rows of data + 1 header
    assert ws.max_row == 3
    # First data row keeps the rank ordering from the input list
    assert ws.cell(row=2, column=1).value == 1
    assert ws.cell(row=2, column=2).value == "TestCouncil/24/0001"
    assert ws.cell(row=3, column=2).value == "TestCouncil/24/0002"
    assert ws.freeze_panes == "C2"
    # 19 base columns + 3 editorial-structure columns (Highlight, Primary
    # cohort, Also in cohorts) + 4 Phase-4 findings columns = 26 → Z.
    # Auto-filter spans all 2 data rows + header.
    assert ws.auto_filter.ref == "A1:Z3"
    findings_header_row = [c.value for c in ws[1]]
    # Editorial-structure headers
    assert findings_header_row[2] == "Highlight"
    assert findings_header_row[3] == "Primary cohort"
    assert findings_header_row[4] == "Also in cohorts"
    # Phase-4 findings columns
    assert "Findings — new disclosures" in findings_header_row
    assert "Findings — refinements" in findings_header_row
    assert "Findings — disclosed MW" in findings_header_row
    assert "Findings — headline" in findings_header_row
    # Humanised lineage column carries the prose, not the raw tag.
    # Column 20 (T) after the cohort columns shifted things right by 3.
    humanised = ws.cell(row=3, column=20).value
    assert "Spatial neighbour" in humanised
    assert "TestCouncil/24/0001" in humanised
    # Last four columns (Findings*) on the data rows should be None — the
    # fixture has no findings attached. Column indices 23-26.
    for r in (2, 3):
        for c in (23, 24, 25, 26):
            assert ws.cell(row=r, column=c).value in (None, "")


def test_write_xlsx_handles_missing_optional_fields(tmp_path):
    """An app row with None / missing description, address, url etc. must not
    raise; the cells just stay empty."""
    minimal = worklist.WorklistData(
        summary={"total": 1, "dc": 1, "adjacent": 0,
                 "unrelated": 0, "unknown": 0, "worklist": 1},
        rows=[{
            "id": 1, "application_ref": "Bare/0001",
            "description": None, "address": None,
            "date_received": None, "url": None,
            "raw_metadata": None, "council_name": None,
            "verdict": "DC", "worth_deep_read": "yes", "confidence": "sure",
            "signals": [], "why": None,
            "discovered_via": [], "foxglove": False,
            "tier1_hits": 0, "storage_hits": 0, "backup_hits": 0,
        }],
        anchors={}, model="fake-granite",
    )
    path = tmp_path / "minimal.xlsx"
    export.write_xlsx(
        path=path, data=minimal, model="fake-granite",
        generated_at=dt.datetime(2026, 5, 15),
    )
    from openpyxl import load_workbook
    wb = load_workbook(path)
    assert wb["Worklist"].max_row == 2  # header + one row
