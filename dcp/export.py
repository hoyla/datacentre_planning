"""Reporter export — produces the Aisha-facing hand-off pair:

  - `worklist_<date>.md`   — top-N narrative markdown, one card per application
  - `worklist_<date>.xlsx` — all worklist entries as a flat sortable / filterable
                             table, suitable for Excel review

Lives under `data/exports/` (gitignored). Re-runnable whenever the universe
or triage changes; downstream consumers see the latest verdict per app via
the shared `dcp.worklist.fetch` query (DISTINCT ON inserted_at).
"""

from __future__ import annotations

import datetime as dt
from pathlib import Path

from dcp import cohorts as cohorts_mod
from dcp import corpus_stats as corpus_stats_mod
from dcp import db, findings as findings_mod, worklist


METHODOLOGY_NOTE = """\
### Methodology in one paragraph

We index UK planning applications from PlanIt (national aggregator, all
councils 2018+) plus the Planning Inspectorate NSIP register, complemented
by a parent-application backfill that walks PlanIt's `associated_id` chain
to recover pre-2018 substantive permissions referenced from procedural
follow-ons. A spatial sweep within 1 km of each DC anchor pulls in
neighbouring applications regardless of description language. A `granite4.1:30b`
LLM (Ollama, local) classifies each application against the rubric in
`data/triage_labelling/rubric.md` — verdict (DC / adjacent / unrelated /
unknown), deep-read recommendation, signals, and confidence. This worklist
filters to `verdict ∈ {{DC, adjacent}} AND deep-read ∈ {{yes, maybe}}`, plus a
safety-net tag for the Foxglove top-10 (which captures procedural follow-ons
the LLM correctly classifies as unrelated but where the family lineage is
still editorially important). Ranking weights primary on-site generation
signals (rubric Tier 1) at 3× and storage signals (rubric Tier 4) at 1×;
backup is a deep-read trigger, not a finding, so it serves only as a
tie-break. Council names follow the current unitary; legacy district names
are preserved in raw metadata. The xlsx companion has every worklist
entry; the markdown carries the top {md_top} curated, head-of-list first.
"""


# ---------------------------------------------------------------------------
# Markdown export
# ---------------------------------------------------------------------------


def _ranked_index_by_ref(rows: list[dict]) -> dict[str, int]:
    """Map each application_ref → its 1-based rank in the full worklist."""
    return {row["application_ref"]: i for i, row in enumerate(rows, 1)}


def _highlights_section(rows_by_ref: dict[str, dict], ranks_by_ref: dict[str, int]) -> list[str]:
    """Editorial highlights — hand-picked headline apps at the top of the
    document. Each links to the full card via its anchor."""
    lines: list[str] = []
    hs = cohorts_mod.highlights()
    if not hs:
        return lines
    lines.append("## Editorial highlights")
    lines.append("")
    lines.append(
        "Hand-picked headline applications. Each is the substantive "
        "anchor of an editorial pattern surfaced by the deep-read pass; "
        "click through to the full card."
    )
    lines.append("")
    for h in hs:
        row = rows_by_ref.get(h.app)
        rank = ranks_by_ref.get(h.app)
        anchor = worklist._ref_anchor(h.app)
        rank_str = f"rank {rank}" if rank else "filtered"
        if row is not None:
            link = f"[`{h.app}`](#{anchor}) ({rank_str})"
        else:
            link = f"`{h.app}` ({rank_str})"
        lines.append(f"- **{link}** — {h.one_liner.strip()}")
        lines.append("")
    return lines


def _cohort_anchor(cohort_name: str) -> str:
    return f"cohort-{cohort_name.replace('_', '-')}"


def _cohort_sections(
    rows_by_ref: dict[str, dict],
    ranks_by_ref: dict[str, int],
    anchors: dict[str, dict],
) -> tuple[list[str], set[str]]:
    """Render every named cohort (in YAML order). Each cohort gets a
    heading, intro paragraph, and the cards for its primary-member apps in
    rank order. Apps whose primary is a different cohort are shown as
    cross-references (one-line, linking to their canonical card).

    Returns the rendered lines + the set of application_refs that landed
    in a cohort section (so the caller can render the remaining apps in
    a separate "Other" section).
    """
    lines: list[str] = []
    rendered_refs: set[str] = set()

    cohorts = cohorts_mod.cohorts()
    if not cohorts:
        return lines, rendered_refs

    lines.append("## Editorial cohorts")
    lines.append("")
    lines.append(
        "Themed groupings of related applications — operator networks, "
        "spatial clusters, planning-route patterns. Within each cohort, "
        "cards run in rank order. Applications that belong to more "
        "than one cohort appear as a full card in their primary cohort "
        "and as a cross-reference in the others."
    )
    lines.append("")

    for cohort in cohorts:
        lines.append(f'<a id="{_cohort_anchor(cohort.name)}"></a>')
        lines.append(f"### {cohort.display_name}")
        lines.append("")
        if cohort.description:
            lines.append(cohort.description)
            lines.append("")
        # Partition apps into "primary here" vs "cross-ref from another cohort".
        primary_apps: list[dict] = []
        cross_refs: list[dict] = []
        for ref in cohort.apps:
            row = rows_by_ref.get(ref)
            if row is None:
                # In universe but filtered out (excluded / duplicate),
                # or not in our DB at all — skip silently. The excluded
                # section at the bottom will show it if relevant.
                continue
            if row.get("primary_cohort") == cohort.name:
                primary_apps.append(row)
            else:
                cross_refs.append(row)

        # Sort each group by rank so the editorially-loudest case in each
        # cohort floats to the top of its section.
        primary_apps.sort(key=lambda r: ranks_by_ref.get(r["application_ref"], 9999))
        cross_refs.sort(key=lambda r: ranks_by_ref.get(r["application_ref"], 9999))

        if not primary_apps and not cross_refs:
            lines.append("*(No applications currently in this cohort.)*")
            lines.append("")
            lines.append("---")
            lines.append("")
            continue

        for row in primary_apps:
            rank = ranks_by_ref[row["application_ref"]]
            lines.append(worklist.render_card(rank, row, anchors, base_level=4))
            rendered_refs.add(row["application_ref"])

        if cross_refs:
            lines.append(f"#### Cross-references to other cohorts")
            lines.append("")
            for row in cross_refs:
                primary = row.get("primary_cohort")
                primary_label = primary
                primary_obj = cohorts_mod.cohort_by_name(primary) if primary else None
                if primary_obj:
                    primary_label = primary_obj.display_name
                ref = row["application_ref"]
                rank = ranks_by_ref.get(ref, "—")
                anchor = worklist._ref_anchor(ref)
                lines.append(
                    f"- [`{ref}`](#{anchor}) — primary cohort is "
                    f"[{primary_label}](#{_cohort_anchor(primary or '')}) "
                    f"(see rank {rank})."
                )
            lines.append("")
        lines.append("---")
        lines.append("")
    return lines, rendered_refs


def _excluded_section(excluded: list[dict]) -> list[str]:
    """The audit-trail list at the bottom: apps filtered out of the
    primary worklist (exclusions + duplicate-tags). Brief one-liners
    only, so reviewers can sanity-check the calls without each one
    eating a full card."""
    lines: list[str] = []
    if not excluded:
        return lines
    lines.append("## Filtered from the worklist")
    lines.append("")
    lines.append(
        "Applications that the deep-read pass has confirmed are NOT data "
        "centres, or are duplicate consultation-copies of substantive "
        "applications elsewhere on this worklist. Shown here so reviewers "
        "can sanity-check (and dispute) the filtering."
    )
    lines.append("")
    # Resolve exclusions YAML to add the notes alongside the tag.
    notes_by_app = {e.app: e.notes for e in cohorts_mod.exclusions()}
    for row in excluded:
        ref = row["application_ref"]
        tags = [t for t in (row.get("discovered_via") or [])
                if t.startswith(("exclude:", "duplicate_of:"))]
        tag_str = ", ".join(f"`{t}`" for t in tags)
        addr = (row.get("address") or "").strip()
        note = notes_by_app.get(ref, "")
        line = f"- **`{ref}`**"
        if tag_str:
            line += f" — {tag_str}"
        if addr:
            line += f" — {addr}"
        lines.append(line)
        if note:
            lines.append(f"  > {note.strip()}")
        lines.append("")
    return lines


def render_markdown(
    *,
    data: worklist.WorklistData,
    top: int,
    model: str,
    generated_at: dt.datetime,
    excluded: list[dict] | None = None,
    stats: dict | None = None,
) -> str:
    s = data.summary
    rows = data.rows
    ranks_by_ref = _ranked_index_by_ref(rows)
    rows_by_ref = {r["application_ref"]: r for r in rows}
    when = generated_at.isoformat(timespec="seconds")

    out: list[str] = []
    out.append(f"# Data-centre planning worklist — {generated_at.date().isoformat()}")
    out.append("")
    out.append(
        f"Investigation: systematic UK DC planning dataset, surfacing on-site "
        f"power-generation signals.  "
        f"Generated {when} from triage model `{model}`."
    )
    out.append("")
    out.append("## At a glance")
    out.append("")
    # Universe line: prefer the corpus-stats variant (date span + by-source
    # breakdown) when available, falling back to the worklist summary alone.
    if stats:
        u = stats["universe"]
        date_span = ""
        if u["date_min"] and u["date_max"]:
            date_span = (
                f" (date received {u['date_min'].isoformat()} → "
                f"{u['date_max'].isoformat()})"
            )
        src_parts = ", ".join(f"{n} from `{name}`" for name, n in u["by_source"])
        out.append(
            f"- **Universe:** {u['total']} applications ingested{date_span}, "
            f"all triaged by `{model}`."
            + (f" Sources: {src_parts}." if src_parts else "")
        )
    else:
        out.append(f"- **Universe:** {s['total']} applications triaged.")
    out.append(
        f"- **Verdict mix:** DC {s['dc']} · adjacent {s['adjacent']} · "
        f"unrelated {s['unrelated']} · unknown {s['unknown']}."
    )
    if stats:
        f = stats["filters"]
        out.append(
            f"- **Worklist size:** {s['worklist']} applications "
            f"(DC/adjacent ∩ deep-read yes/maybe, or Foxglove-tagged) — "
            f"{f['excluded']} confirmed-not-a-DC after deep-read and "
            f"{f['duplicates']} consultation-stage duplicates filtered."
        )
        d = stats["documents"]
        out.append(
            f"- **Documents on file:** {d['docs_total']} fetched, "
            f"covering {d['apps_with_docs']} applications."
        )
        fn = stats["findings"]
        out.append(
            f"- **Document-extracted findings:** {fn['findings_total']} findings on "
            f"{fn['apps_with_findings']} applications from "
            f"{fn['documents_with_findings']} documents — an editorially-"
            f"selected sample, not the full worklist (see *How to read "
            f"this* below)."
        )
    else:
        out.append(
            f"- **Worklist size:** {s['worklist']} applications "
            f"(DC/adjacent ∩ deep-read yes/maybe, or Foxglove-tagged); "
            f"deep-read false-positives and known duplicates filtered."
        )
    out.append(f"- **Companion xlsx:** all {s['worklist']} worklist entries, flat table for filtering.")
    out.append("")

    out.extend(_highlights_section(rows_by_ref, ranks_by_ref))

    out.append(METHODOLOGY_NOTE.format(md_top=min(top, len(rows))))
    out.append("")
    out.append("### How to read each card")
    out.append("")
    out.append(
        "Each card shows the LLM's verdict and confidence, the rubric-tiered signal "
        "counts, the council and address, the signals extracted from the description, "
        "a one-line model reasoning, the full description verbatim (so you can "
        "sanity-check the LLM), a `Why this is on the worklist` explanation of how "
        "the application entered our universe, document-extracted findings (when the "
        "deep-read pass has run on this application), and a link to the source-portal "
        "record. Substantive deep-read claims should be drillable back to those documents."
    )
    out.append("")
    out.append("---")
    out.append("")

    # Cohort sections — themed grouping of related apps. Returns the set
    # of refs already rendered so the "Other applications" section can
    # cover only the rest.
    cohort_lines, rendered_refs = _cohort_sections(rows_by_ref, ranks_by_ref, data.anchors)
    out.extend(cohort_lines)

    # Other applications: everything not already in a cohort, capped at
    # the top-`top` so the document doesn't balloon. Apps in cohorts have
    # already been shown.
    other_rows = [r for r in rows if r["application_ref"] not in rendered_refs]
    other_rows = other_rows[:top]
    if other_rows:
        out.append("## Other applications")
        out.append("")
        out.append(
            f"Top {len(other_rows)} ranked applications that aren't currently "
            "assigned to a named cohort. The xlsx companion has the full "
            f"{s['worklist']}-entry worklist for sortable / filterable review."
        )
        out.append("")
        for row in other_rows:
            rank = ranks_by_ref[row["application_ref"]]
            out.append(worklist.render_card(rank, row, data.anchors, base_level=3))
            out.append("---")
            out.append("")

    if excluded:
        out.extend(_excluded_section(excluded))

    return "\n".join(out)


# ---------------------------------------------------------------------------
# XLSX export
# ---------------------------------------------------------------------------


def write_xlsx(
    *,
    path: Path,
    data: worklist.WorklistData,
    model: str,
    generated_at: dt.datetime,
    excluded: list[dict] | None = None,
) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Worklist"

    # Highlights are flagged via the per-app one-liner — build the set
    # once so the row loop can mark each highlighted app.
    highlight_refs = {h.app for h in cohorts_mod.highlights()}

    headers = [
        "Rank",
        "Application ref",
        "Highlight",
        "Primary cohort",
        "Also in cohorts",
        "Verdict",
        "Deep read recommended",
        "Confidence",
        "Tier-1 hits",
        "Storage hits",
        "Backup hits",
        "Foxglove top-10",
        "Council",
        "Address",
        "Date received",
        "Application type",
        "Signals",
        "LLM reasoning",
        "Discovered via (raw)",
        "Discovered via (humanised)",
        "Source portal URL",
        "Description",
        # Phase-4 columns (NULL/blank for apps without findings yet).
        "Findings — new disclosures",
        "Findings — refinements",
        "Findings — disclosed MW",
        "Findings — headline",
    ]
    ws.append(headers)
    for col_idx, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="DDDDDD")
        cell.alignment = Alignment(vertical="center", wrap_text=False)

    for rank, row in enumerate(data.rows, 1):
        descr = row.get("description") or ""
        signals = ", ".join(row.get("signals") or [])
        via_raw = ", ".join(row.get("discovered_via") or [])
        via_human = "\n".join(
            f"• {line}" for line in worklist.expand_lineage(
                row.get("discovered_via"), data.anchors,
            )
        )
        date_received = row.get("date_received")
        date_str = date_received.isoformat() if date_received else ""
        council = (
            row.get("council_name")
            or (row.get("raw_metadata") or {}).get("area_name")
            or ""
        )
        app_type = (row.get("raw_metadata") or {}).get("app_type") or ""
        # Phase-4 columns: populated only when the app has findings rows.
        row_findings = row.get("findings") or []
        f_counts = findings_mod.category_counts(row_findings)
        new_count = f_counts.get(findings_mod.CATEGORY_NEW) or None
        ref_count = f_counts.get(findings_mod.CATEGORY_REFINEMENT) or None
        disclosed_mw = findings_mod.disclosed_mw_total(row_findings)
        headline = findings_mod.headline_disclosure(row_findings) or None
        # Editorial-structure cells: highlight flag + cohort assignments.
        primary_cohort_name = row.get("primary_cohort") or ""
        primary_cohort_obj = (
            cohorts_mod.cohort_by_name(primary_cohort_name)
            if primary_cohort_name else None
        )
        primary_cohort_label = primary_cohort_obj.display_name if primary_cohort_obj else primary_cohort_name
        also_in: list[str] = []
        for c_name in (row.get("also_in_cohorts") or []):
            c_obj = cohorts_mod.cohort_by_name(c_name)
            also_in.append(c_obj.display_name if c_obj else c_name)
        also_in_str = ", ".join(also_in)
        ws.append([
            rank,
            row["application_ref"],
            "yes" if row["application_ref"] in highlight_refs else "",
            primary_cohort_label,
            also_in_str,
            row["verdict"],
            row["worth_deep_read"],
            row["confidence"],
            row["tier1_hits"],
            row["storage_hits"],
            row["backup_hits"],
            "yes" if row.get("foxglove") else "",
            council,
            row.get("address") or "",
            date_str,
            app_type,
            signals,
            row.get("why") or "",
            via_raw,
            via_human,
            row.get("url") or "",
            descr,
            new_count,
            ref_count,
            disclosed_mw,
            headline,
        ])

    # Column widths — calibrated for the dominant content per column.
    # After inserting Highlight + cohort columns at C/D/E, the original
    # columns shift by 3. Letter assignments below reflect post-insertion
    # layout: A Rank, B Ref, C Highlight, D Primary cohort, E Also-in,
    # F Verdict … through Z Findings-headline.
    widths = {
        "A": 6,   # Rank
        "B": 36,  # Application ref
        "C": 10,  # Highlight
        "D": 26,  # Primary cohort
        "E": 28,  # Also in cohorts
        "F": 10, "G": 9, "H": 10,                  # Verdict / DR / Conf
        "I": 8,  "J": 8, "K": 9,                   # Tier-1 / Storage / Backup
        "L": 9,                                     # Foxglove
        "M": 28,                                    # Council
        "N": 50,                                    # Address
        "O": 13,                                    # Date received
        "P": 14,                                    # App type
        "Q": 50,                                    # Signals
        "R": 60,                                    # LLM reasoning
        "S": 40, "T": 60,                          # Discovered via raw / humanised
        "U": 50,                                    # Source portal URL
        "V": 80,                                    # Description
        "W": 11, "X": 11, "Y": 12, "Z": 80,         # Findings columns
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # Wrap text on the long columns so the row stays usable when sorted.
    wrap_cols = {"Q", "R", "S", "T", "V", "Z"}
    last_row = len(data.rows) + 1
    for col in wrap_cols:
        for r in range(2, last_row + 1):
            ws[f"{col}{r}"].alignment = Alignment(vertical="top", wrap_text=True)
    # Top-align everything for sortable readability
    for r in range(2, last_row + 1):
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=r, column=c)
            if cell.alignment.wrap_text is None:
                cell.alignment = Alignment(vertical="top")

    ws.freeze_panes = "C2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{last_row}"

    # Filtered apps sheet — the false-positives + duplicates, kept
    # accessible so reviewers can sanity-check (or dispute) the calls
    # without each one eating a row on the main Worklist tab.
    if excluded:
        notes_by_app = {e.app: e.notes for e in cohorts_mod.exclusions()}
        filt = wb.create_sheet("Filtered")
        filt_headers = [
            "Application ref",
            "Filter tag(s)",
            "Address",
            "Verdict",
            "Deep read recommended",
            "Source portal URL",
            "Notes",
        ]
        filt.append(filt_headers)
        for col_idx in range(1, len(filt_headers) + 1):
            cell = filt.cell(row=1, column=col_idx)
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="DDDDDD")
            cell.alignment = Alignment(vertical="center")
        for row in excluded:
            ref = row["application_ref"]
            tags = ", ".join(
                t for t in (row.get("discovered_via") or [])
                if t.startswith(("exclude:", "duplicate_of:"))
            )
            filt.append([
                ref,
                tags,
                row.get("address") or "",
                row.get("verdict") or "",
                row.get("worth_deep_read") or "",
                row.get("url") or "",
                notes_by_app.get(ref, ""),
            ])
        for col, w in {"A": 36, "B": 50, "C": 50, "D": 10, "E": 11,
                       "F": 50, "G": 70}.items():
            filt.column_dimensions[col].width = w
        for r in range(2, filt.max_row + 1):
            for col_letter in ("B", "C", "F", "G"):
                filt[f"{col_letter}{r}"].alignment = Alignment(
                    vertical="top", wrap_text=True,
                )
        filt.freeze_panes = "B2"
        filt.auto_filter.ref = f"A1:{get_column_letter(len(filt_headers))}{filt.max_row}"

    # Methodology / summary on a second sheet
    meta = wb.create_sheet("Methodology")
    meta_rows = [
        ("Generated at", generated_at.isoformat(timespec="seconds")),
        ("Triage model", model),
        ("Total triaged", data.summary["total"]),
        ("DC", data.summary["dc"]),
        ("Adjacent", data.summary["adjacent"]),
        ("Unrelated", data.summary["unrelated"]),
        ("Unknown", data.summary["unknown"]),
        ("Worklist size", data.summary["worklist"]),
        ("", ""),
        ("Methodology", METHODOLOGY_NOTE.format(md_top="(see markdown)").strip()),
    ]
    for r, (k, v) in enumerate(meta_rows, 1):
        meta.cell(row=r, column=1, value=k).font = Font(bold=True)
        meta.cell(row=r, column=2, value=v).alignment = Alignment(
            vertical="top", wrap_text=True,
        )
    meta.column_dimensions["A"].width = 22
    meta.column_dimensions["B"].width = 110

    wb.save(path)


# ---------------------------------------------------------------------------
# Top-level entry point
# ---------------------------------------------------------------------------


def export_worklist(
    *,
    model: str = "granite4.1:30b",
    output_dir: Path,
    md_top: int = 50,
    generated_at: dt.datetime | None = None,
    md_path: Path | None = None,
    xlsx_path: Path | None = None,
) -> dict[str, Path]:
    """Generate the markdown + xlsx pair for the current worklist.

    Markdown carries the top `md_top` cards for narrative reading; the xlsx
    has every worklist entry, ranked, with both the raw and humanised
    `discovered_via` lineage columns so Aisha can filter however she likes.

    Default path naming is `worklist_<date>.{md,xlsx}` inside `output_dir`.
    `md_path` / `xlsx_path` override the defaults for callers that need
    a specific naming convention (e.g. the release-folder orchestrator).
    """
    output_dir.mkdir(parents=True, exist_ok=True)
    generated_at = generated_at or dt.datetime.now()
    today = generated_at.date().isoformat()
    md_path = md_path or (output_dir / f"worklist_{today}.md")
    xlsx_path = xlsx_path or (output_dir / f"worklist_{today}.xlsx")

    with db.connect() as conn:
        data = worklist.fetch(conn, model=model)  # no limit — full worklist
        excluded = worklist.fetch_excluded(conn, model=model)
        stats = corpus_stats_mod.collect(conn, model=model)

    md_path.write_text(render_markdown(
        data=data, top=md_top, model=model, generated_at=generated_at,
        excluded=excluded, stats=stats,
    ))
    write_xlsx(
        path=xlsx_path, data=data, model=model,
        generated_at=generated_at, excluded=excluded,
    )
    return {"markdown": md_path, "xlsx": xlsx_path}
